#!/usr/bin/env python3
"""
extract_pool.py — Canonical data-pool extractor for the equity-research engine.

WHY THIS EXISTS
  Capital IQ / NSE / broker exports frequently ship a *multi-tab* workbook in a
  single file (e.g. one `EstimatesReport.xls` whose tabs are Consensus / Recent
  Changes / Multiples / Surprise / Trends / Revisions). Legacy `.xls` is OLE2/BIFF
  and `.xlsx` cells are binary — the values are NOT recoverable with grep/strings,
  and a filename-only inventory sees ONE opaque file and silently drops every tab
  but the first. This splits every workbook into one readable text extract PER TAB,
  and extracts pdf/rtf to text, so the front-door triage and every downstream
  specialist read EVERY tab — nothing left behind (CLAUDE.md §2 reuse, §11 data
  sufficiency, §24 "leave no data behind").

READS EVERY FORMAT. Workbooks (every tab), PDFs (text layer), RTF/Word, HTML/MHTML, and — when the
text layer is absent — image-only/scanned PDFs AND standalone images (screenshots, photographed
tables, charts). Those visual sources are transcribed to Markdown by Claude vision (best; verbatim
tables/charts) when ANTHROPIC_API_KEY is set, else by tesseract OCR (offline floor). Results are cached
per (file + method + model), so the cost — and any token spend — is paid once per file, ever.

ANY LANGUAGE, TRANSCRIBED VERBATIM. Non-English filings (Arabic, Mandarin, Japanese, …) are extracted
exactly like English ones — the vision/OCR transcriber reproduces the ORIGINAL script verbatim and
never translates (a filing's numbers and labels stay as filed). Translation into English is a
reading-layer job the analyst agents do when they read these extracts, NOT an extraction step — so a
foreign-language document is fully in the pool, never a data gap (CLAUDE.md §27). Only a genuine
EXTRACT-FAIL row (corrupt/encrypted/illegible) is missing data.

SINGLE SOURCE OF TRUTH. Consumers:
  - Layer-0 `*-data-triage` agents  -> run at ingestion, list every tab as a row
  - commands/research/verify-evidence.md -> post-hoc corpus build (--corpus)
  - ui/server/src/data-status.ts     -> cockpit tab listing (--list-json)
  - ui/server/src/readiness.ts       -> pre-flight gate (--readiness-json; complete offline read, vision OFF)

USAGE
  python3 extract_pool.py <DATA_PATH> <OUT_DIR> [--force] [--corpus PATH] [--vision|--no-vision]
  python3 extract_pool.py --list-json <FILE>     # print one file's sheets as JSON; no writes
  # Vision (image-only PDFs + images) defaults to EXTRACT_VISION (=claude); the pre-flight gate forces
  # it off and completes the provider-neutral local read without a pool deadline. Env: EXTRACT_VISION, EXTRACT_VISION_MODEL,
  # EXTRACT_OCR_MAX_PAGES, EXTRACT_OCR_DPI, EXTRACT_OCR_FILE_TIMEOUT_S, EXTRACT_OCR_BUDGET_S.

OUTPUTS (in OUT_DIR)
  <stem>__<sheet>.txt   one per spreadsheet tab (header + TSV body)
  <stem>.txt            one per pdf/rtf (extracted text)
  manifest.json         machine-readable inventory (per source, per tab)
  manifest.md           human-readable table to paste into the triage inventory

CONTRACT
  Tolerant   — missing libs / unreadable files become EXTRACT-FAIL rows; never aborts.
  Idempotent — skips when manifest.json is newer than every source AND this script,
               unless --force.
  Pure-ish   — only writes inside OUT_DIR (keeps the Google Drive pool pristine).
"""
import sys
import os
import re
import io
import errno
import fcntl
import json
import glob
import shutil
import base64
import hashlib
import stat
import subprocess
import tempfile
from contextlib import contextmanager
import urllib.request
import urllib.error
import time as _time  # NB: `from datetime import ... time` below would shadow a bare `import time`
import email as _email
import html as _htmlmod
from datetime import datetime, date, time, timezone

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(os.path.dirname(TOOLS_DIR))
SCRIPTS_DIR = os.path.join(REPO_ROOT, "scripts")
CONNECTOR_REGISTRY_ROOT = os.path.join(REPO_ROOT, ".claude", "connectors")

WORKBOOK_EXTS = {"xls", "xlsx", "xlsm"}
TEXT_EXTS = {"txt", "md", "csv", "tsv"}
PDF_EXTS = {"pdf"}
RTF_EXTS = {"rtf"}
# Standalone images — screenshots (IBKR/Bloomberg), photographed tables, charts. Read via vision/OCR.
IMAGE_EXTS = {"png", "jpg", "jpeg", "gif", "webp", "tif", "tiff", "bmp"}
# Google Drive pointer stubs — tiny JSON, no real content
POINTER_EXTS = {"gdoc", "gsheet", "gslides"}


def _decision_time_utc(now=None):
    """Capture one aware UTC instant for every release-eligibility decision."""
    if now is None:
        return datetime.now(timezone.utc)
    if not isinstance(now, datetime) or now.tzinfo is None or now.utcoffset() is None:
        raise ValueError("extraction now must be a timezone-aware datetime")
    return now.astimezone(timezone.utc)


# ---------- stable reads on a cloud-synced pool ----------
# The data pool is normally a Google Drive for Desktop MIRROR folder (see GDRIVE in ui/server/src/
# config.ts). Drive — and macOS itself, via the `com.apple.lastuseddate#PS` attribute — writes
# EXTENDED ATTRIBUTES onto pool files independently of their content. An xattr write bumps
# `st_ctime_ns` and NOTHING else: size, mtime and the inode identity are untouched, and the file's
# bytes are byte-for-byte identical.
#
# The readers below used to fold `st_ctime_ns` into an identity tuple compared before and after the
# read, and rejected ANY delta as "changed during read". A Drive sync pass that landed while the
# extractor was reading therefore rejected files whose content never moved — the INDIAMART incident,
# where one sync pass turned all 87 pool files into `unsafe-input` / "pool input rejected" rows and
# left the readiness gate with zero usable evidence for a company whose pool was perfectly fine.
#
# What actually has to hold is that the bytes returned are a COMPLETE, CONSISTENT snapshot of the
# pinned inode. `st_size` + `st_mtime_ns` + (dev, ino, nlink) carry exactly that: a content write —
# including an in-place same-size byte flip — always bumps mtime. A ctime-only delta does not imply
# a content change, so instead of failing it we RE-READ through the same pinned descriptor and
# compare digests, which PROVES the bytes rather than inferring it from a clock.
#
# KNOWN RESIDUAL, stated plainly rather than papered over. Excluding ctime gives up exactly one
# detection: a same-uid local writer who rewrites the file mid-read with SAME-LENGTH bytes and then
# restores st_mtime_ns to the nanosecond with utimensat(). ctime is the one stamp no syscall can
# roll back, so it was the only metadata evidence of that sequence — and it is precisely the stamp a
# cloud-sync client makes unusable. Nothing else regresses: a torn read, a differing-length rewrite,
# a rewrite that does not restore mtime, a path/inode swap, a hard-link alias and a symlink are all
# still caught, and the digest re-read additionally catches a rewrite that lands after the read.
# The compensating control for evidence that carries provenance is `_snapshot_pool_file`'s
# attested-digest binding, which fails closed on exactly this substitution. For an ordinary pool
# document there is no provenance to forge, and the pool is the user's own Drive folder — so the
# trade is: absorb the failure mode that empties a real evidence pool weekly, in exchange for a
# same-uid attack that could equally have written the file a moment before the read began.
#
# Anything that still looks like a real mid-read change — or a transient cloud-filesystem error
# (Drive hydrating a dehydrated file surfaces as ETIMEDOUT / "[Errno 60] Operation timed out") — is
# RETRIED rather than failed. A file is only rejected once the retry budget is spent, so a sync pass
# costs a few hundred milliseconds instead of the whole run's evidence.
STABLE_READ_ATTEMPTS = max(1, int(os.environ.get("EXTRACT_STABLE_READ_ATTEMPTS") or 4))
STABLE_READ_BACKOFF_S = max(0.0, float(os.environ.get("EXTRACT_STABLE_READ_BACKOFF_S") or 0.15))
STABLE_READ_BACKOFF_MAX_S = 2.0

# Errors a network/cloud-backed filesystem raises transiently while hydrating or re-syncing a file.
# EINTR/EAGAIN are ordinary interruptions; ENOENT/ESTALE cover Drive replacing a file by
# unlink+rename mid-pass (the path is re-pinned from scratch on the retry). A genuinely absent file
# simply exhausts the budget and fails, as it should.
_TRANSIENT_READ_ERRNOS = frozenset(
    getattr(errno, name) for name in
    ("ETIMEDOUT", "EIO", "EINTR", "EAGAIN", "EWOULDBLOCK", "EBUSY", "ENOENT", "ESTALE",
     "ENXIO", "EDEADLK", "ENOTCONN", "EHOSTDOWN")
    if hasattr(errno, name)
)


class _UnstableRead(ValueError):
    """A read that did not settle. RETRYABLE.

    Subclasses ValueError so every existing `except (OSError, ValueError)` call site still catches
    it once the retry budget is spent — no caller changes, same fail-closed contract at the end.
    A refusal that is NOT retryable (symlink, hard-link alias, non-canonical path) stays a plain
    ValueError and is never retried: those are verdicts about the file, not races.
    """


class _ReadRace(_UnstableRead):
    """The pin was lost BEFORE any bytes were read (root/dir/inode swapped under us).

    Nothing was parsed, so re-pinning from scratch is always safe and always correct.
    """


class _ContentMoved(_UnstableRead):
    """The file's own bytes moved DURING the read, so what we assembled may be torn.

    Whether this is retryable depends on what the file IS, which is why `_stable_read` takes
    `retry_content_change`:

      • A POOL input is the user's evidence file on a Google Drive mirror. Drive genuinely
        re-downloads and rewrites it. Re-reading and returning a clean snapshot of the settled
        file is both safe and what the user wants — and when those bytes carry sealed provenance,
        `_snapshot_pool_file` still re-hashes them against the attested digest and fails closed.
      • A CONNECTOR PROJECTION is a content-addressed sealed object. Its invariant is stricter:
        a mid-read content change is rejected outright, never re-read, so an attacker cannot pair
        later bytes with an earlier-verified digest.
    """


def _content_identity(info):
    """The stat fields that move if and only if the file's BYTES can have moved.

    Deliberately EXCLUDES st_ctime_ns: ctime tracks inode-metadata changes (xattrs, permissions,
    link count), which a cloud-sync client writes constantly without touching a single byte.
    """
    return (info.st_dev, info.st_ino, info.st_size, info.st_mtime_ns, info.st_nlink)


def _reread_digest_matches(fd, size, expected_digest):
    """Re-read the SAME pinned descriptor and confirm the bytes still hash to `expected_digest`.

    Used when only ctime moved during a read. Reading through the already-open fd cannot be
    redirected by a path swap, so this compares the identical inode we just parsed — it proves a
    metadata-only touch instead of trusting a timestamp.
    """
    os.lseek(fd, 0, os.SEEK_SET)
    digest = hashlib.sha256()
    remaining = size
    while remaining:
        chunk = os.read(fd, min(remaining, 1024 * 1024))
        if not chunk:
            return False
        digest.update(chunk)
        remaining -= len(chunk)
    return not os.read(fd, 1) and digest.hexdigest() == expected_digest


def _stable_read(read_once, retry_content_change):
    """Run one pinned read attempt, retrying an unsettled read or a transient cloud-FS error.

    `retry_content_change` decides whether a mid-read CONTENT change is re-read (pool inputs) or
    rejected outright (content-addressed connector projections) — see `_ContentMoved`.
    """
    last = None
    for attempt in range(STABLE_READ_ATTEMPTS):
        try:
            return read_once()
        except _ContentMoved as exc:
            if not retry_content_change:
                raise
            last = exc
        except _ReadRace as exc:
            last = exc
        except OSError as exc:
            if exc.errno not in _TRANSIENT_READ_ERRNOS:
                raise
            last = exc
        if attempt + 1 < STABLE_READ_ATTEMPTS and STABLE_READ_BACKOFF_S:
            _time.sleep(min(STABLE_READ_BACKOFF_S * (2 ** attempt), STABLE_READ_BACKOFF_MAX_S))
    raise last


def _read_regular_nofollow_bytes_once(path):
    before = os.lstat(path)
    if (stat.S_ISLNK(before.st_mode) or not stat.S_ISREG(before.st_mode)
            or before.st_nlink != 1):
        raise ValueError("connector projection must be a unique regular non-symlink file")
    fd = os.open(path, os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0))
    try:
        opened = os.fstat(fd)
        identity = _content_identity(opened)
        ctime_before = opened.st_ctime_ns
        if opened.st_nlink != 1 or identity != _content_identity(before):
            raise _ReadRace("connector projection changed before read")
        chunks = []
        remaining = opened.st_size
        while remaining:
            chunk = os.read(fd, min(remaining, 1024 * 1024))
            if not chunk:
                raise _ContentMoved("connector projection truncated during read")
            chunks.append(chunk)
            remaining -= len(chunk)
        if os.read(fd, 1):
            raise _ContentMoved("connector projection grew during read")
        after_open = os.fstat(fd)
        after = os.lstat(path)
        if identity != _content_identity(after_open) or identity != _content_identity(after):
            raise _ContentMoved("connector projection changed during read")
        raw = b"".join(chunks)
        if ctime_before != after_open.st_ctime_ns or ctime_before != after.st_ctime_ns:
            # metadata-only touch (xattr / chmod / cloud-sync): prove the bytes, don't trust the clock
            if not _reread_digest_matches(fd, opened.st_size, hashlib.sha256(raw).hexdigest()):
                raise _ContentMoved("connector projection changed during read")
        return raw
    finally:
        os.close(fd)


def _read_regular_nofollow_bytes(path):
    # sealed, content-addressed object: ride out metadata churn and transient IO, but a real
    # mid-read content change stays a hard refusal (never re-read under an earlier digest).
    return _stable_read(lambda: _read_regular_nofollow_bytes_once(path), retry_content_change=False)


def _pool_regular_descriptors(data_path, rel):
    """Open one pool-relative file without trusting any descendant symlink.

    The pool root itself may be a configured Drive symlink, so it is resolved
    once. Every component below it is then opened relative to pinned directory
    descriptors with ``O_NOFOLLOW``. The returned parent descriptor lets the
    caller prove the filename still names the same inode after reading.
    """
    rel = str(rel).replace("/", os.sep)
    components = rel.split(os.sep)
    if (not components or any(part in {"", ".", ".."} for part in components)):
        raise ValueError("pool-relative path is not canonical")
    root = os.path.realpath(os.path.abspath(data_path))
    root_info = os.lstat(root)
    if stat.S_ISLNK(root_info.st_mode) or not stat.S_ISDIR(root_info.st_mode):
        raise ValueError("pool root is not a real directory")
    descriptors = []
    try:
        directory_fd = os.open(
            root, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0) | getattr(os, "O_NOFOLLOW", 0),
        )
        descriptors.append(directory_fd)
        opened_root = os.fstat(directory_fd)
        if (opened_root.st_dev, opened_root.st_ino) != (root_info.st_dev, root_info.st_ino):
            # a race, not a verdict about the file — _stable_read re-pins from scratch
            raise _ReadRace("pool root changed before read")
        for component in components[:-1]:
            info = os.stat(component, dir_fd=directory_fd, follow_symlinks=False)
            if stat.S_ISLNK(info.st_mode) or not stat.S_ISDIR(info.st_mode):
                raise ValueError("pool path contains a non-directory or symlink component")
            next_fd = os.open(
                component,
                os.O_RDONLY | getattr(os, "O_DIRECTORY", 0) | getattr(os, "O_NOFOLLOW", 0),
                dir_fd=directory_fd,
            )
            opened = os.fstat(next_fd)
            if (opened.st_dev, opened.st_ino) != (info.st_dev, info.st_ino):
                os.close(next_fd)
                raise _ReadRace("pool directory changed before read")
            descriptors.append(next_fd)
            directory_fd = next_fd
        name = components[-1]
        before = os.stat(name, dir_fd=directory_fd, follow_symlinks=False)
        if (stat.S_ISLNK(before.st_mode) or not stat.S_ISREG(before.st_mode)
                or before.st_nlink != 1):
            raise ValueError("pool input must be a unique regular non-symlink file")
        file_fd = os.open(
            name, os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0), dir_fd=directory_fd,
        )
        opened = os.fstat(file_fd)
        if opened.st_nlink != 1:
            # a writable hard-link alias is a VERDICT about the file, not a race — never retry it
            os.close(file_fd)
            raise ValueError("pool input must be a unique regular non-symlink file")
        if (opened.st_dev, opened.st_ino) != (before.st_dev, before.st_ino):
            os.close(file_fd)
            raise _ReadRace("pool input changed before read")
        return descriptors, directory_fd, name, file_fd, opened
    except BaseException:
        for descriptor in reversed(descriptors):
            os.close(descriptor)
        raise


def _read_pool_regular_bytes_once(data_path, rel):
    descriptors, parent_fd, name, file_fd, opened = _pool_regular_descriptors(data_path, rel)
    identity = _content_identity(opened)
    ctime_before = opened.st_ctime_ns
    try:
        chunks = []
        remaining = opened.st_size
        while remaining:
            chunk = os.read(file_fd, min(remaining, 1024 * 1024))
            if not chunk:
                raise _ContentMoved("pool input truncated during read")
            chunks.append(chunk)
            remaining -= len(chunk)
        if os.read(file_fd, 1):
            raise _ContentMoved("pool input grew during read")
        after_open = os.fstat(file_fd)
        after_path = os.stat(name, dir_fd=parent_fd, follow_symlinks=False)
        if identity != _content_identity(after_open) or identity != _content_identity(after_path):
            raise _ContentMoved("pool input changed during read")
        raw = b"".join(chunks)
        if ctime_before != after_open.st_ctime_ns or ctime_before != after_path.st_ctime_ns:
            # Only ctime moved: a metadata-only touch — Google Drive writing its
            # `com.google.drivefs.item-id` xattr, macOS stamping `com.apple.lastuseddate#PS`, a
            # chmod. None of those change a byte. Prove that by re-reading the SAME pinned
            # descriptor and comparing digests, rather than failing the file on a timestamp.
            if not _reread_digest_matches(file_fd, opened.st_size, hashlib.sha256(raw).hexdigest()):
                raise _ContentMoved("pool input changed during read")
        return raw
    finally:
        os.close(file_fd)
        for descriptor in reversed(descriptors):
            os.close(descriptor)


def _read_pool_regular_bytes(data_path, rel):
    # user evidence on a cloud-synced mirror: a genuine re-sync is re-read to a settled snapshot.
    # Sealed provenance is still enforced downstream by _snapshot_pool_file's attested-digest check.
    return _stable_read(lambda: _read_pool_regular_bytes_once(data_path, rel), retry_content_change=True)


def _snapshot_pool_file(data_path, rel, snapshot_pool, attested_sha256=None):
    """Snapshot one pool file for parsing, optionally bound to an attested digest.

    Full-pool extraction snapshots every input first and subsequently parses
    both evidence and provenance from those same captured bytes.  The optional
    digest remains useful to direct callers that already hold an attestation.
    """
    raw = _read_pool_regular_bytes(data_path, rel)
    if attested_sha256 is not None:
        actual = hashlib.sha256(raw).hexdigest()
        if actual != attested_sha256:
            raise ValueError(
                "projection bytes changed after provenance verification "
                f"(attested {attested_sha256!r}, snapshot {actual})"
            )
    target = os.path.join(snapshot_pool, *str(rel).replace("/", os.sep).split(os.sep))
    os.makedirs(os.path.dirname(target), exist_ok=True)
    with open(target, "wb") as fh:
        fh.write(raw)
    return target

# ---------- external-data provenance (frameworks/EXTERNAL_DATA.md) ----------
# Externally ingested documents live under data/<TICKER>/external/<provider>/ with a
# `<file>.source.json` provenance sidecar (provider, source_type, §4 tier, as-of, license).
# The sidecar is METADATA, never a document: it is skipped as a source and folded into its
# document's manifest row as `provenance`. These predicates run at CONSUMPTION sites only —
# never inside iter_pool_files, which also feeds is_fresh (filtering there would blind the
# freshness check to sidecar edits and external files).
SIDECAR_SUFFIX = ".source.json"
METHODOLOGY_ONLY_FIELD = "methodology_only"
METHODOLOGY_ONLY_ERROR = "methodology-only WILTW source excluded from runtime evidence"
METHODOLOGY_ONLY_SNIFF_CHARS = 60_000


def _methodology_words(value):
    """Compatibility wrapper around the canonical publication-boundary rule."""
    if SCRIPTS_DIR not in sys.path:
        sys.path.insert(0, SCRIPTS_DIR)
    from connector_contract import _methodology_words as shared_methodology_words
    return shared_methodology_words(value)


def methodology_only_source_reason(name, text="", provenance=None):
    """Return the same identity verdict used before canonical publication."""
    if SCRIPTS_DIR not in sys.path:
        sys.path.insert(0, SCRIPTS_DIR)
    from connector_contract import methodology_only_source_reason as shared_reason
    return shared_reason(name, text, provenance)


def _is_sidecar(name):
    return name.lower().endswith(SIDECAR_SUFFIX)


def _is_external_rel(rel):
    """True when a pool-relative path sits under an `external/` folder segment."""
    return "external" in rel.replace(os.sep, "/").split("/")[:-1]


def _v2_projection_manifest(data_path, rel):
    """Return the v2 manifest which owns this subject projection path, if any."""
    if SCRIPTS_DIR not in sys.path:
        sys.path.insert(0, SCRIPTS_DIR)
    from connector_contract import load_valid_manifests  # lazy: extractor remains standalone

    subject = os.path.basename(os.path.realpath(data_path))
    manifests, _skipped = load_valid_manifests(CONNECTOR_REGISTRY_ROOT)
    owners = []
    rel_posix = rel.replace(os.sep, "/")
    for manifest in manifests:
        if manifest.get("manifest_version") != 2 or subject not in manifest.get("subjects", []):
            continue
        template = manifest["output_path"].replace("<SUBJECT>", subject)
        prefix = f"data/{subject}/"
        if template.startswith(prefix):
            template = template[len(prefix):]
        pattern = "^" + re.escape(template).replace(re.escape("<as_of>"), r"\d{4}-\d{2}-\d{2}") + "$"
        if re.fullmatch(pattern, rel_posix):
            owners.append(manifest)
    if len(owners) > 1:
        raise ValueError("projection path has multiple v2 connector owners")
    return owners[0] if owners else None


def _verify_v2_projection(
        data_path, rel, parsed, manifest, decision_at, *, canonical_data_path=None):
    """Return ``current``, ``expired``, ``disputed``, or ``superseded`` for an exact projection."""
    if SCRIPTS_DIR not in sys.path:
        sys.path.insert(0, SCRIPTS_DIR)
    if TOOLS_DIR not in sys.path:
        sys.path.insert(0, TOOLS_DIR)
    from connector_contract import canonical_json_bytes, connector_storage_paths
    from connector_vintages import vintage_is_eligible_at
    import run_connectors

    # `data_path` may be the private byte snapshot used by one extraction
    # transaction.  Canonical connector heads/blobs remain rooted beside the
    # real subject pool, while the materialized projection and sidecar are
    # deliberately read only from that immutable snapshot.
    pool_root = os.path.realpath(data_path)
    canonical_pool_root = os.path.realpath(canonical_data_path or data_path)
    subject = os.path.basename(canonical_pool_root)
    data_root = os.path.dirname(canonical_pool_root)
    rel_posix = rel.replace(os.sep, "/")
    expected_legacy_path = f"{subject}/{rel_posix}"
    if parsed.get("connector_id") != manifest["id"]:
        raise ValueError("sidecar connector_id does not match projection owner")
    doc = os.path.join(pool_root, rel)
    projected_bytes = _read_pool_regular_bytes(data_path, rel)
    projected_payload = json.loads(projected_bytes.decode("utf-8"))
    if canonical_json_bytes(projected_payload) != projected_bytes:
        raise ValueError("projection payload bytes are not canonical")
    projected_hash = hashlib.sha256(projected_bytes).hexdigest()
    if projected_hash != parsed.get("content_sha256"):
        raise ValueError("projection payload does not match sidecar content_sha256")
    sidecar_bytes = _read_pool_regular_bytes(data_path, rel + SIDECAR_SUFFIX)
    if canonical_json_bytes(parsed) != sidecar_bytes:
        raise ValueError("projection sidecar bytes are not canonical")

    current_path = connector_storage_paths(data_root, manifest, subject)["current"]
    current_bytes = _read_regular_nofollow_bytes(current_path)
    current = json.loads(current_bytes.decode("utf-8"))
    if canonical_json_bytes(current) != current_bytes:
        raise ValueError("canonical current pointer bytes are not canonical")
    is_current_claim = (
        parsed.get("vintage_id") == current.get("vintage_id")
        or current.get("legacy_path") == expected_legacy_path
    )
    if is_current_claim:
        run_connectors.verify_current_pointer(manifest, subject, data_root, current)
        if canonical_json_bytes(run_connectors._payload_from_current(data_root, current)) != projected_bytes:
            raise ValueError("current projection payload bytes differ from sealed blob")
        if current.get("legacy_path") != expected_legacy_path:
            raise ValueError("current projection path does not match canonical legacy_path")
        if not isinstance(current.get("provenance"), dict):
            raise ValueError("canonical current has no provenance")
        if canonical_json_bytes(parsed) != canonical_json_bytes(current["provenance"]):
            raise ValueError("projection sidecar does not exactly match canonical current provenance")
        if projected_hash != current.get("content_sha256"):
            raise ValueError("current projection payload hash differs from canonical current")
        if not vintage_is_eligible_at(current, decision_at):
            return "expired"
        # Integrity and expiry are not the whole admission test. When a primary and a fallback publish
        # different payloads for the same release, run_connectors records provider_disagreement only AFTER
        # both vintages and projections are already published, and read_point_in_time then answers
        # usable: false for the series. Without re-asking at decision_at, extraction would hand research
        # both conflicting projections as citable present-tense evidence — each individually well-formed.
        if run_connectors._point_in_time_disagreement(data_root, manifest, subject, decision_at):
            return "disputed"
        return "current"

    vintage_id = parsed.get("vintage_id")
    resolved = run_connectors.resolve_vintage_id(
        data_root, manifest["series_id"], subject, vintage_id,
    )
    if not isinstance(resolved, dict) or resolved.get("usable") is not True:
        raise ValueError("projection does not resolve to one committed canonical vintage")
    vintage = resolved.get("vintage")
    if (not isinstance(vintage, dict)
            or vintage.get("connector_id") != manifest["id"]
            or vintage.get("dataset_id") != manifest["dataset_id"]
            or vintage.get("series_id") != manifest["series_id"]
            or vintage.get("subject") != subject):
        raise ValueError("resolved vintage identity does not match projection owner")
    if vintage.get("legacy_path") != expected_legacy_path:
        raise ValueError("superseded projection path does not match committed legacy_path")
    if not isinstance(vintage.get("provenance"), dict):
        raise ValueError("resolved vintage has no provenance")
    if canonical_json_bytes(parsed) != canonical_json_bytes(vintage["provenance"]):
        raise ValueError("projection sidecar does not exactly match committed vintage provenance")
    if projected_hash != vintage.get("content_sha256"):
        raise ValueError("superseded projection payload hash differs from committed vintage")
    if canonical_json_bytes(projected_payload) != canonical_json_bytes(resolved.get("payload")):
        raise ValueError("superseded projection payload differs from committed blob")
    return "superseded"


def _collect_sidecars(data_path, decision_at, inventory=None, *, canonical_data_path=None):
    """Return provenance parsed from one exact pool snapshot.

    ``data_path`` and ``inventory`` identify the bytes being extracted.  When
    they point at a private snapshot, ``canonical_data_path`` keeps connector
    registry/current lookups anchored beside the real subject pool without
    reopening either the live projected document or its live sidecar.
    """
    out = {}
    inventory = tuple(inventory) if inventory is not None else _complete_pool_inventory(data_path)
    for p in inventory:
        base = os.path.basename(p)
        if not _is_sidecar(base):
            continue
        try:
            rel_sidecar = os.path.relpath(p, data_path)
            parsed = json.loads(_read_pool_regular_bytes(data_path, rel_sidecar).decode("utf-8"))
        except Exception:  # noqa — ordinary malformed sidecars degrade; owned v2 projections are caught below
            continue
        # A sidecar that parses to a NON-object (a bare JSON array/scalar) is malformed the same way an
        # unparseable one is: skip it so the doc falls to path-derived provenance, never a non-dict stored
        # in prov_map that the fold would later call .get() on (AttributeError). Only objects are provenance.
        if isinstance(parsed, dict):
            # Internal verification state is earned below, never accepted from
            # an untrusted sidecar that happens to use the same field name.
            parsed = {
                key: value for key, value in parsed.items()
                if key not in {"_verified_content_sha256", "_verified_snapshot_sha256"}
            }
            rel = os.path.relpath(p, data_path)[: -len(SIDECAR_SUFFIX)]
            try:
                owner = _v2_projection_manifest(canonical_data_path or data_path, rel)
            except Exception as exc:
                owner = None
                parsed = {**parsed, "_integrity_error": f"bad connector projection: {exc}"}
            claims_v2 = (parsed.get("manifest_version") == 2 or "vintage_id" in parsed
                         or "publisher_contract_fingerprint" in parsed)
            if owner is None and claims_v2 and "_integrity_error" not in parsed:
                parsed = {**parsed, "_integrity_error": "bad connector projection: no valid v2 manifest owns this path"}
            if owner is None and parsed.get("routed_by") == "ingest_external.py" \
                    and "_integrity_error" not in parsed:
                expected_raw = parsed.get("sha256")
                doc = os.path.realpath(os.path.join(os.path.realpath(data_path), rel))
                try:
                    if (not isinstance(expected_raw, str)
                            or not re.fullmatch(r"[0-9a-f]{64}", expected_raw)):
                        raise ValueError("ordinary ingest sidecar lacks a valid raw sha256")
                    projected_bytes = _read_pool_regular_bytes(data_path, rel)
                    if hashlib.sha256(projected_bytes).hexdigest() != expected_raw:
                        raise ValueError("ordinary ingest sha256 does not match routed bytes")
                    parsed = {**parsed, "_ordinary_ingest_hash_verified": True,
                              "_verified_content_sha256": expected_raw}
                except Exception as exc:
                    parsed = {**parsed, "_integrity_error": f"bad routed external input: {exc}"}
            elif owner is None and "sha256" in parsed and parsed.get("routed_by") != "ingest_external.py":
                # A hand-written or connector sidecar cannot self-assert a raw
                # hash. Only the router contract above makes this field trusted.
                parsed = {key: value for key, value in parsed.items() if key != "sha256"}
            expected = parsed.get("content_sha256")
            if owner is not None and "_integrity_error" not in parsed:
                try:
                    projection_state = _verify_v2_projection(
                        data_path, rel, parsed, owner, decision_at,
                        canonical_data_path=canonical_data_path,
                    )
                    # Carry the digest this verification actually attested, so the later
                    # _snapshot_pool_file read can be bound to THESE bytes rather than trusting that
                    # the projection did not change in between (see _snapshot_pool_file's re-check).
                    parsed = {**parsed, "_projection_state": projection_state,
                              "_verified_content_sha256": parsed.get("content_sha256")}
                except Exception as exc:
                    parsed = {**parsed, "_integrity_error": f"bad connector projection: {exc}"}
            elif expected is not None and "_integrity_error" not in parsed:
                doc = os.path.realpath(os.path.join(os.path.realpath(data_path), rel))
                root = os.path.realpath(data_path)
                try:
                    if not (doc == root or doc.startswith(root + os.sep)):
                        raise ValueError("projected document path escapes the pool")
                    projected_bytes = _read_pool_regular_bytes(data_path, rel)
                    payload = json.loads(projected_bytes.decode("utf-8"))
                    if SCRIPTS_DIR not in sys.path:
                        sys.path.insert(0, SCRIPTS_DIR)
                    from connector_contract import canonical_json_bytes
                    if canonical_json_bytes(payload) != projected_bytes:
                        raise ValueError("projected document bytes are not canonical")
                    actual = hashlib.sha256(projected_bytes).hexdigest()
                    if not isinstance(expected, str) or actual != expected:
                        raise ValueError(f"content_sha256 mismatch (expected {expected!r}, got {actual})")
                    parsed = {**parsed, "_verified_content_sha256": actual}
                except Exception as exc:
                    parsed = {**parsed, "_integrity_error": f"bad connector projection: {exc}"}
            out[rel] = parsed

    # An ordinary hand-dropped external document may omit its sidecar and fall back to conservative,
    # path-derived provenance. A v2 connector projection may not: its payload and sidecar are one materialized
    # view of canonical current, so accepting the payload alone would turn a broken canonical projection into
    # an unrelated tier-9 document. Scan documents after sidecars so missing, invalid-JSON, and non-object
    # sidecars all fail closed when (and only when) a valid v2 manifest owns that exact projection path.
    for p in inventory:
        base = os.path.basename(p)
        if _is_sidecar(base):
            continue
        rel = os.path.relpath(p, data_path)
        if not _is_external_rel(rel) or rel in out:
            continue
        try:
            owner = _v2_projection_manifest(canonical_data_path or data_path, rel)
        except Exception as exc:
            out[rel] = {"_integrity_error": f"bad connector projection: {exc}"}
            continue
        if owner is not None:
            sidecar_path = p + SIDECAR_SUFFIX
            reason = "missing sidecar" if not os.path.exists(sidecar_path) else "sidecar is unreadable or not a JSON object"
            out[rel] = {"_integrity_error": f"bad connector projection: {reason}"}
    return out


# §4/§5 masquerade guard (EXTERNAL_DATA.md §4). Each source_type earns a §4 tier CEILING (lower = more
# trusted). A sidecar may declare a MORE conservative tier, never a more trusted one — a scrape / channel
# check / broker note can never fold into the pool stamped as a tier-5 vendor number. Enforced HERE, at fold
# time, so no downstream specialist ever sees an over-claimed tier no matter who wrote the sidecar (a hand
# drop, or an auto-built connector's fetcher whose self-reported tier is not to be trusted).
SOURCE_TYPE_MAX_TRUST = {
    "alt_data_panel": 5, "vendor_export": 5, "paid_api": 5, "official_data": 5,
    "peer_transcript": 6,  # a competitor's own earnings-call transcript — tier 6 about the PEER (§4);
                           # Level-1 inference about the SUBJECT (competitive-intel module owns its use)
    "broker_research": 7,
    "expert_call": 9, "channel_check": 9, "management_meeting": 9,
    "external_other": 9,
}


def _enforce_tier_ceiling(prov):
    """Clamp a sidecar's §4 tier DOWN to the ceiling its source_type permits (never up); fill a missing/
    non-numeric tier from the source_type; flag any over-claim as `tier_corrected` so the manifest + triage
    see the correction. An UNKNOWN / missing / typo'd source_type fails CLOSED to the conservative
    external_other floor (tier 9 — the same default a no-sidecar drop gets): a self-declared tier we cannot
    classify is never trusted verbatim. (Falling through untouched here was a fail-OPEN hole — a sidecar with
    an off-list source_type and a low tier folded in at that trusted tier, a worse masquerade than the tier-5
    one this gate blocks.) Mutates and returns the provenance dict."""
    st = prov.get("source_type")
    # Unknown / missing source_type → the most conservative ceiling (external_other, tier 9), never a bypass.
    ceiling = SOURCE_TYPE_MAX_TRUST.get(st, SOURCE_TYPE_MAX_TRUST["external_other"])
    declared = prov.get("tier")
    if isinstance(declared, bool) or not isinstance(declared, (int, float)):
        prov["tier"] = ceiling  # derive from the ceiling when absent / non-numeric (bool is not a tier)
    elif declared < ceiling:  # over-claimed a more-trusted tier than the source_type earns → clamp + flag
        prov["tier"] = ceiling
        prov["tier_corrected"] = {"declared": declared, "enforced": ceiling,
                                  "reason": f"source_type '{st}' may be cited at most at §4 tier {ceiling}"}
    return prov


def _finish_row(row, rel, prov_map):
    """Enrich a manifest source row with its pool-relative path (when nested — duplicate
    basenames across subfolders stay distinguishable) and, for external documents, the
    `external` flag + `provenance` from the sidecar. A direct drop WITHOUT a sidecar still
    gets path-derived provenance — provider = the containing folder's name, conservative
    tier-9 default (EXTERNAL_DATA.md §2 `external_other`) — so triage always has the
    provider/tier labels its citations require, never `unknown provider · unclassified`."""
    base = row.get("file", "")
    if rel and rel != base:
        row["path"] = rel.replace(os.sep, "/")
    if rel and _is_external_rel(rel):
        row["external"] = True
        prov = prov_map.get(rel)
        if isinstance(prov, dict) and prov.get("_integrity_error"):
            # Do not disguise an unusable connector projection as an ordinary sidecar-less tier-9 drop.
            # The source row already carries status=fail + the exact error; this small marker makes the
            # provenance state explicit without assigning it a citable source type or evidence tier.
            row["provenance"] = {"integrity_status": "failed", "usable": False}
            return row
        if isinstance(prov, dict) and prov.get("_projection_state") == "superseded":
            # Retain the immutable dated projection for audit/replay, but do
            # not expose its provider fields as present-tense citable evidence.
            row["provenance"] = {"integrity_status": "superseded", "usable": False}
            return row
        if isinstance(prov, dict) and prov.get("_projection_state") == "disputed":
            # Two providers disagree for this release: the series has no single answer, so nothing here is
            # citable until an operator resolves it. Visible for diagnosis, never usable as evidence.
            row["provenance"] = {"integrity_status": "disputed", "usable": False}
            return row
        if isinstance(prov, dict) and prov.get("_projection_state") == "expired":
            # Expiry is an expected release-clock state, not broken integrity.
            # Keep the row visible for diagnosis without exposing citable
            # provider fields or letting readiness count it as current data.
            row["provenance"] = {"eligibility_status": "expired", "usable": False}
            return row
        if not prov:
            parts = rel.replace(os.sep, "/").split("/")
            folder = parts[-2] if len(parts) >= 2 else ""
            prov = {"provider": folder if folder and folder != "external" else "unfiled",
                    "source_type": "external_other", "tier": 9,
                    "provenance_basis": "path-derived (no sidecar)"}
        row["provenance"] = _enforce_tier_ceiling({k: v for k, v in prov.items() if not k.startswith("_")})
    return row


def _prov_summary(row):
    """One-line human summary for manifest.md: 'external: YipitData · alt_data_panel · tier 5 · as-of 2026-03-31'."""
    if not row.get("external"):
        return ""
    p = row.get("provenance") or {}
    bits = [str(p.get("provider") or "unknown provider"), str(p.get("source_type") or "unclassified")]
    if p.get("tier"):
        bits.append(f"tier {p['tier']}")
    tc = p.get("tier_corrected")
    if isinstance(tc, dict):  # only a real correction dict (written by _enforce_tier_ceiling) renders a flag;
        bits.append(f"⚠ tier corrected (declared {tc.get('declared')}, over-claimed)")  # a pre-declared non-dict is ignored
    if p.get("as_of"):
        bits.append(f"as-of {p['as_of']}")
    return "external: " + " · ".join(bits)


def _ensure_deps():
    """The extractor needs xlrd/openpyxl for workbooks, pypdf for portable PDF text, and
    striprtf for portable CIQ relationship-table extraction.
    On a PEP-668 "externally-managed" system Python (e.g. Homebrew) you cannot
    pip-install them globally, so the engine ships an isolated venv at
    `.claude/tools/.venv` (gitignored; recreate via setup-tools.sh). If the current
    interpreter lacks the libs but that venv has them, transparently re-exec under it
    so `python3 extract_pool.py ...` just works regardless of which python invoked it.
    [fix F-EXTRACT-DEP — validated on the CRM pool: Homebrew py3.14 had neither lib,
    so every .xls silently degraded to useless 'fallback-text'.]"""
    try:
        import xlrd  # noqa
        import openpyxl  # noqa
        import pypdf  # noqa  [DD-16] pure-Python PDF text fallback when poppler is absent
        from striprtf.striprtf import rtf_to_text  # noqa
        return
    except Exception:
        pass
    # A venv's bin/python is a SYMLINK to the same base binary, so comparing
    # realpath(executable) can't tell venv from base — use an env sentinel instead,
    # which also guarantees we never re-exec more than once (no loop).
    if os.environ.get("_EXTRACT_POOL_VENV") == "1":
        return
    here = os.path.dirname(os.path.abspath(__file__))
    venv_py = os.path.join(here, ".venv", "bin", "python")
    if os.path.exists(venv_py):
        os.environ["_EXTRACT_POOL_VENV"] = "1"  # invoke via the symlink path so the venv activates
        os.execv(venv_py, [venv_py, os.path.abspath(__file__)] + sys.argv[1:])
    # No venv: don't abort — workbook reads will mark 'missing-dependency' LOUDLY
    # (a real failure with the install command), never silent fallback-text.


def sniff_format(path):
    """Return the TRUE format from magic bytes / head, because Capital IQ mislabels
    extensions wholesale — a `.doc` that is MHTML, a `.rtf` that is binary Word, an
    `.xls` that is an HTML table. Content wins over extension.
    [fix F-EXTRACT — live-validated on the CRM pool, where all three mismatches occurred.]
    Returns one of: ole2 | zip | pdf | rtf | mime | html | text | empty | unreadable."""
    try:
        with open(path, "rb") as fh:
            head = fh.read(4096)
    except Exception:
        return "unreadable"
    if not head.strip():
        return "empty"
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "ole2"          # OLE2 Compound Document — BIFF .xls OR binary Word .doc
    if head[:4] == b"PK\x03\x04":
        return "zip"           # zip container — .xlsx / .docx
    if head[:5] == b"%PDF-":
        return "pdf"
    # standalone images (screenshots / photographed tables / charts) — read by vision or OCR.
    # Unambiguous magics only (full TIFF magic, not bare "II"/"MM"); BMP/odd images are caught by
    # extension in the extract loop, so a text file starting "BM"/"II" is never misread as an image.
    if (head[:8] == b"\x89PNG\r\n\x1a\n" or head[:3] == b"\xff\xd8\xff" or head[:4] == b"GIF8"
            or (head[:4] == b"RIFF" and head[8:12] == b"WEBP")
            or head[:4] in (b"II*\x00", b"MM\x00*")):
        return "image"
    if head[:5].lower() == b"{\\rtf":
        return "rtf"
    low = head[:1024].lower()
    if (b"mime-version:" in low or low.lstrip().startswith(b"x-sender:")
            or b"content-type: multipart" in low or low.lstrip().startswith(b"from:")):
        return "mime"          # MHTML (HTML wrapped in MIME) — CIQ filing exports
    if b"<html" in low or b"<table" in low or b"<!doctype html" in low or b"<spreadsheet" in low:
        return "html"          # HTML table mislabeled .xls
    return "text"


def _fmt(v):
    """Render a cell value greppably: integral floats lose the '.0', dates -> ISO."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        r = repr(v)
        # [fix F25] repr() emits scientific notation for very large/small floats, which defeats the
        # verify-evidence / resolve_citations grep (a cited "1500000000" can't match "1.5e+09"). Expand it.
        return f"{v:.12f}".rstrip("0").rstrip(".") if ("e" in r or "E" in r) else r
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    return str(v)


def _sanitize(name, fallback="sheet"):
    s = re.sub(r"[^A-Za-z0-9._-]+", "-", str(name)).strip("-._")
    s = re.sub(r"-{2,}", "-", s)
    return s or fallback


def _unique(used, base):
    name = base
    i = 2
    while name in used:
        name = f"{base}-{i}"
        i += 1
    used.add(name)
    return name


# ---------- workbook readers: return list of (sheet_name, rows, cols, list_of_rows) ----------

def _read_bytes_retry(path, tries=5):
    """Read a file's bytes with bounded retry on a transient OSError. [fix F02]
    On a Google Drive FUSE mount, concurrent access to the same .xls (the 6 module
    triage agents race on identical files) raises 'OSError: Resource deadlock avoided'
    (errno 11) — a RECOVERABLE lock that the old code turned into a permanent extraction
    failure. Reading into memory here also lets the workbook readers avoid mmap entirely."""
    last = None
    for i in range(tries):
        try:
            with open(path, "rb") as fh:
                return fh.read()
        except OSError as e:  # EAGAIN / EDEADLK / transient FUSE lock
            last = e
            _time.sleep(0.2 * (i + 1))
    raise last


def _detect_units(rows):
    """Scan a tab's first rows for a declared unit/currency header (CIQ tabs carry
    'Currency: …' / 'Units: …' / 'in millions' near the top). [fix F04] Surfaced into
    the manifest so downstream reads carry scale/currency and a crore-vs-million mix
    can be caught instead of silently producing a 10x error."""
    head = " ".join(c for r in rows[:20] for c in r if c)[:1200]  # CIQ pads empty rows before the header block
    hits = []
    for pat in (r"currency\s*[:=]\s*[A-Za-z ]+", r"units?\s*[:=]\s*[A-Za-z0-9 ()]+",
                r"\bin (?:thousands|millions|billions|lakhs?|crores?)\b",
                r"\b(?:USD|INR|EUR|GBP|JPY)\b", r"[₹$€£]\s*in\s*\w+"):
        m = re.search(pat, head, re.I)
        if m:
            hits.append(m.group(0).strip())
    return "; ".join(dict.fromkeys(hits))[:160]  # dedupe, cap length


def _read_xls(path):
    import xlrd  # xlrd >= 2.0 is purpose-built for legacy .xls
    # xlrd logs warnings to stdout by default — redirect so stdout stays clean
    # (the UI parses --list-json stdout as JSON, and corpus consumers grep it).
    # [fix F02] open from in-memory bytes (file_contents=) so xlrd never mmaps the file
    # over the FUSE mount, the source of the 'Resource deadlock avoided' failures.
    wb = xlrd.open_workbook(file_contents=_read_bytes_retry(path), on_demand=True, logfile=sys.stderr)
    datemode = wb.datemode
    out = []
    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        rows = []
        for r in range(sh.nrows):
            cells = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    val = bool(val)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    val = None
                cells.append(_fmt(val))
            rows.append(cells)
        out.append((name, sh.nrows, sh.ncols, rows))
        wb.unload_sheet(name)
    return out


def _read_xlsx(path):
    import openpyxl
    # [fix F02] load from in-memory bytes (retry-guarded) rather than the FUSE path.
    wb = openpyxl.load_workbook(io.BytesIO(_read_bytes_retry(path)), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        # [fix F24; corrected per PR#9 review] read_only mode trusts the stored <dimension> tag,
        # which some CIQ/broker exporters write stale/missing — causing iter_rows to silently stop
        # early and drop trailing rows/cols while the tab still looks 'ok'. Force a real cell scan.
        # NB: reset_dimensions is a METHOD in read-only mode — it must be CALLED, not assigned. The
        # old `ws.reset_dimensions = True` overwrote the method and never ran, so the truncation it
        # was meant to fix kept reproducing.
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
        rows = []
        maxc = 0
        for row in ws.iter_rows(values_only=True):
            cells = [_fmt(v) for v in row]
            # trim trailing empties for a tidy extract
            while cells and cells[-1] == "":
                cells.pop()
            maxc = max(maxc, len(cells))
            rows.append(cells)
        out.append((ws.title, len(rows), maxc, rows))
    wb.close()
    return out


def read_workbook(path, ext):
    return _read_xls(path) if ext == "xls" else _read_xlsx(path)


def _nonempty_cells(rows):
    return sum(1 for r in rows for c in r if c.strip())


def _tab_text(src_name, sheet_name, idx, total, rows, ncols):
    buf = io.StringIO()
    buf.write(f"# SOURCE: {src_name}\n")
    buf.write(f"# SHEET: {sheet_name}  ({idx} of {total})\n")
    buf.write(f"# DIMS: {len(rows)} rows x {ncols} cols\n")
    buf.write("# ---\n")
    for r in rows:
        buf.write("\t".join(r) + "\n")
    return buf.getvalue()


# ---------- pdf / rtf ----------

def _read_pdf_text(path, max_pages=None):
    # pdftotext (poppler) preferred; -layout keeps tables aligned. max_pages caps
    # the scan for fast sniffing (a cover page already says "Annual Report" / "10-K").
    # Falls back to pure-Python pypdf (auto-bootstrapped in the venv, no system dep) when
    # poppler is absent or yields nothing, so PDF extraction never hard-depends on a system
    # binary. [fix DD-16 — live-validated on the TMCV pool, where Homebrew lacked pdftotext
    # and all 8 statutory PDFs (annual reports + transcripts) produced zero extracts.]
    try:
        cmd = ["pdftotext"] + (["-l", str(max_pages)] if max_pages else []) + ["-layout", path, "-"]
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout, None
        # poppler present but produced no text (image-only scan, or a soft error) — try pypdf
        txt, err = _read_pdf_py(path, max_pages)
        return (txt, None) if txt.strip() else ("", err or f"pdftotext rc={r.returncode}; pypdf produced no text")
    except FileNotFoundError:
        # poppler not installed at all — pure-Python fallback
        txt, err = _read_pdf_py(path, max_pages)
        return (txt, None) if txt.strip() else ("", err or "pdftotext not installed and pypdf produced no text")
    except Exception as e:  # noqa
        return "", f"{type(e).__name__}: {e}"


def _read_pdf_py(path, max_pages=None):
    """Pure-Python PDF text fallback via pypdf (no system dependency). Returns (text, error).
    Poppler's `pdftotext -layout` is preferred for table alignment; this keeps the engine
    working on machines without poppler. An image-only / scanned PDF still yields no text here
    (the caller marks it image-only) — that is a real limitation, not a fallback failure."""
    try:
        import pypdf
    except Exception as e:
        return "", f"pypdf unavailable ({e}); run .claude/tools/setup-tools.sh"
    try:
        reader = pypdf.PdfReader(path)
        pages = reader.pages[:max_pages] if max_pages else reader.pages
        out = []
        for pg in pages:
            try:
                out.append(pg.extract_text() or "")
            except Exception:
                continue  # one bad page shouldn't sink the whole document
        return "\n".join(out), None
    except Exception as e:
        return "", f"pypdf: {type(e).__name__}: {e}"


# ---------- visual reading for image-only PDFs + standalone images ----------
# A scanned PDF (e.g. a Capital IQ "Preliminary Report" saved as page images) and a standalone
# image (a screenshot, a photographed table, a chart) carry NO text layer — pdftotext/pypdf yield
# nothing and a filename inventory drops them. Rather than leave that data unread, TRANSCRIBE it:
#   1. Claude vision (best; verbatim Markdown incl. tables/charts) when ANTHROPIC_API_KEY is set,
#   2. else tesseract OCR (offline floor; garbles complex tables but better than nothing).
# The result is cached ONCE per (file identity + method + model) under .ocr-cache, so a --force
# re-check and every module's run reuse it — the cost (and any token spend) is paid once, ever.
# Vision is OFF in readiness (it must not spend tokens before the user proceeds). Readiness gives
# local OCR no pool-wide deadline and freezes both raw bytes and every completed local result, so
# later children reuse that exact provider-neutral generation rather than upgrading it in-run.

def _env_int(name, default):
    try:
        v = int(os.environ.get(name, "") or "")
        return v if v > 0 else default
    except Exception:
        return default

_OCR_ENABLED = os.environ.get("EXTRACT_OCR", "1") != "0"
_OCR_MAX_PAGES = _env_int("EXTRACT_OCR_MAX_PAGES", 30)   # first N pages carry the statements/notes we cite
_OCR_DPI = _env_int("EXTRACT_OCR_DPI", 200)              # 200 dpi is ample for machine text, ~2x faster than 300
_OCR_PER_FILE_TIMEOUT = _env_int("EXTRACT_OCR_FILE_TIMEOUT_S", 300)
# Optional pool-wide visual-read wall-clock budget for non-admission callers. Readiness explicitly
# overrides it to 0 (unbounded) because a frozen generation may not defer work to later children.
_OCR_POOL_BUDGET_S = _env_int("EXTRACT_OCR_BUDGET_S", 0)
_OCR_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".ocr-cache")
_ocr_deadline = None  # monotonic deadline for the pool budget; set per extract_pool() call

# --- vision (Claude) config ---
# EXTRACT_VISION: "claude" (default) reads image-only PDFs + images with Claude vision when a key is
# present; "off"/"0"/"none" uses tesseract only. Model is tunable — default to the strongest; set
# EXTRACT_VISION_MODEL=claude-sonnet-5 (or -haiku) to cut cost, since transcription barely differs by tier.
_VISION_PROVIDER = os.environ.get("EXTRACT_VISION", "claude").strip().lower()
_VISION_ON_ENV = _VISION_PROVIDER not in ("", "0", "off", "false", "none", "no")
_VISION_MODEL = os.environ.get("EXTRACT_VISION_MODEL", "claude-opus-4-8").strip()
_VISION_MAX_TOKENS = _env_int("EXTRACT_VISION_MAX_TOKENS", 8000)
_VISION_BATCH_PAGES = _env_int("EXTRACT_VISION_BATCH_PAGES", 4)  # images per Messages request
_ANTHROPIC_BASE = os.environ.get("ANTHROPIC_BASE_URL", "https://api.anthropic.com").rstrip("/")
_ANTHROPIC_VERSION = "2023-06-01"
_vision_on = False  # set per extract_pool() call (OFF in the readiness pre-flight; ON in-run)

_VISION_SYSTEM = (
    "You are a meticulous financial-document transcriber. Transcribe the page image(s) to clean "
    "GitHub-flavored Markdown, VERBATIM. Reproduce every number, label, date, currency, unit, and "
    "footnote EXACTLY as shown — never round, infer, translate, or correct. Render every table as a "
    "Markdown table, preserving rows, columns, headers, and units. Preserve reading order and headings. "
    "For a chart or graph, state its title, axes, and units, then read off each series as a small table. "
    "Do NOT summarize, interpret, analyze, or add any commentary. If a value is illegible write "
    "[illegible]. Output ONLY the Markdown transcription."
)


def _anthropic_key():
    return os.environ.get("ANTHROPIC_API_KEY") or ""


def _vision_configured():
    """Vision requested for this pass AND a key is present."""
    return _vision_on and _VISION_PROVIDER == "claude" and bool(_anthropic_key())


def _visual_source_identity(path):
    """Content identity shared by live-pool paths and their private snapshots.

    Extraction parses a securely opened snapshot, never the mutable source
    pathname. A path/mtime cache key would therefore hide an OCR result written
    for the source from the byte-identical snapshot. Hashing one stable,
    no-follow read makes both names converge without reopening an unverified
    path in the parser.

    The stability comparison EXCLUDES st_ctime_ns on purpose. The returned identity IS a sha256 of
    the bytes, so a metadata-only touch (Google Drive writing its item-id xattr, macOS stamping
    lastuseddate) cannot change it. Folding ctime in made every such touch fall through to the
    `unreadable:<path>` fallback, which scopes the key to a path and therefore MISSES the vision
    cache — silently re-paying Claude vision token spend for a file whose bytes never moved, against
    this module's own promise that the cost is "paid once per file, ever".
    """
    descriptor = None
    try:
        before = os.lstat(path)
        if stat.S_ISLNK(before.st_mode) or not stat.S_ISREG(before.st_mode):
            raise ValueError("visual source is not a regular file")
        descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0))
        opened = os.fstat(descriptor)
        expected = (
            before.st_dev, before.st_ino, before.st_size, before.st_mtime_ns,
        )
        if expected != (
            opened.st_dev, opened.st_ino, opened.st_size, opened.st_mtime_ns,
        ):
            raise ValueError("visual source changed before cache identity read")
        digest = hashlib.sha256()
        remaining = opened.st_size
        while remaining:
            chunk = os.read(descriptor, min(remaining, 1024 * 1024))
            if not chunk:
                raise ValueError("visual source truncated during cache identity read")
            digest.update(chunk)
            remaining -= len(chunk)
        if os.read(descriptor, 1):
            raise ValueError("visual source grew during cache identity read")
        after_open = os.fstat(descriptor)
        after_path = os.lstat(path)
        observed_open = (
            after_open.st_dev, after_open.st_ino, after_open.st_size, after_open.st_mtime_ns,
        )
        observed_path = (
            after_path.st_dev, after_path.st_ino, after_path.st_size, after_path.st_mtime_ns,
        )
        if expected != observed_open or expected != observed_path:
            raise ValueError("visual source changed during cache identity read")
        return f"sha256:{digest.hexdigest()}|size:{opened.st_size}"
    except Exception:
        # The caller will fail the actual read separately. A path-scoped key is NOT enough: it is
        # stable across runs, so two DIFFERENT documents that ever occupy the same pathname share
        # one cache entry — replace `investor-presentation.pdf` with the next quarter's deck and,
        # if the identity read fails both times (a freshly-synced Drive placeholder raising
        # ETIMEDOUT is exactly that case), the Q1 transcription is served as the Q2 deck's text.
        # That is a §20 bad-extraction with no visible symptom. A per-call nonce makes an unreadable
        # source MISS the cache always, so a failed identity can only cost a re-read, never
        # attribute one document's content to another.
        return f"unreadable:{os.path.abspath(path)}|{os.urandom(8).hex()}"
    finally:
        if descriptor is not None:
            os.close(descriptor)


def _read_cache_path(path, tag):
    """Cache key = stable content identity + method/model tag."""
    ident = _visual_source_identity(path)
    try:
        os.makedirs(_OCR_CACHE_DIR, exist_ok=True)
    except Exception:
        return None
    return os.path.join(_OCR_CACHE_DIR, hashlib.sha1(f"{tag}|{ident}".encode()).hexdigest() + ".txt")


def _cache_read(cache):
    if cache and os.path.exists(cache):
        try:
            t = open(cache, encoding="utf-8", errors="ignore").read()
            if t.strip():
                return t
        except Exception:
            pass
    return None


def _cache_write(cache, txt):
    if cache:
        try:
            _atomic_text(cache, txt)
        except Exception:
            pass


def _ocr_available():
    """pdftoppm (poppler) + tesseract present? Returns True if page-OCR can run."""
    return bool(shutil.which("pdftoppm") and shutil.which("tesseract"))


def _render_pdf_pages(path, out_dir):
    """Rasterise up to _OCR_MAX_PAGES pages to PNG via pdftoppm (poppler). Shared by vision + OCR."""
    prefix = os.path.join(out_dir, "pg")
    subprocess.run(
        ["pdftoppm", "-png", "-r", str(_OCR_DPI), "-f", "1", "-l", str(_OCR_MAX_PAGES), path, prefix],
        capture_output=True, timeout=_OCR_PER_FILE_TIMEOUT,
    )
    return sorted(glob.glob(prefix + "*.png"))


def _tesseract_pngs(pngs):
    """OCR a list of pre-rendered PNGs, bounded by one wall-clock deadline across the loop."""
    deadline = _time.monotonic() + _OCR_PER_FILE_TIMEOUT
    parts = []
    for png in pngs:
        remaining = deadline - _time.monotonic()
        if remaining < 3:
            break  # out of time — keep what we OCR'd; the rest come on the next (cached) pass
        try:
            r = subprocess.run(["tesseract", png, "stdout", "--psm", "1"],
                               capture_output=True, text=True, timeout=remaining)
        except subprocess.TimeoutExpired:
            break
        if r.returncode == 0 and r.stdout.strip():
            parts.append(r.stdout)
    return "\n".join(parts)


def _normalize_image_to_png(path, out_dir):
    """Best-effort: convert/downscale any image to a vision-friendly PNG (<=1600px). Uses macOS `sips`
    or ImageMagick if present; else returns the original when it's already a web-supported type."""
    dst = os.path.join(out_dir, "img.png")
    if shutil.which("sips"):
        r = subprocess.run(["sips", "-s", "format", "png", "-Z", "1600", path, "--out", dst],
                           capture_output=True, timeout=60)
        if r.returncode == 0 and os.path.exists(dst):
            return dst
    conv = "magick" if shutil.which("magick") else ("convert" if shutil.which("convert") else "")
    if conv:
        r = subprocess.run([conv, path, "-resize", "1600x1600>", dst], capture_output=True, timeout=60)
        if r.returncode == 0 and os.path.exists(dst):
            return dst
    head = b""
    try:
        head = open(path, "rb").read(16)
    except Exception:
        pass
    if head[:8] == b"\x89PNG\r\n\x1a\n" or head[:3] == b"\xff\xd8\xff" or head[:4] == b"GIF8" \
            or (head[:4] == b"RIFF"):
        return path  # already a type the vision API accepts; send as-is
    return ""


def _vision_transcribe(png_paths, label=""):
    """Transcribe rendered PNGs to Markdown via the Anthropic Messages API (vision), in page batches.
    Returns text; raises on any API/network error so the caller can fall back to tesseract."""
    key = _anthropic_key()
    if not key or not png_paths:
        return ""
    out = []
    for i in range(0, len(png_paths), max(1, _VISION_BATCH_PAGES)):
        batch = png_paths[i:i + max(1, _VISION_BATCH_PAGES)]
        content = []
        for j, pp in enumerate(batch):
            try:
                b = open(pp, "rb").read()
            except Exception:
                continue
            content.append({"type": "image", "source": {"type": "base64", "media_type": "image/png",
                            "data": base64.b64encode(b).decode("ascii")}})
            content.append({"type": "text", "text": f"— page {i + j + 1} —"})
        if not content:
            continue
        content.append({"type": "text", "text": "Transcribe the page(s) above to Markdown, verbatim."})
        body = json.dumps({
            "model": _VISION_MODEL, "max_tokens": _VISION_MAX_TOKENS,
            "system": _VISION_SYSTEM,
            "messages": [{"role": "user", "content": content}],
        }).encode()
        req = urllib.request.Request(
            f"{_ANTHROPIC_BASE}/v1/messages", data=body, method="POST",
            headers={"content-type": "application/json", "x-api-key": key,
                     "anthropic-version": _ANTHROPIC_VERSION})
        try:
            with urllib.request.urlopen(req, timeout=_OCR_PER_FILE_TIMEOUT) as resp:
                data = json.loads(resp.read().decode("utf-8", "ignore"))
        except urllib.error.HTTPError as e:
            detail = ""
            try:
                detail = e.read().decode("utf-8", "ignore")[:200]
            except Exception:
                pass
            raise RuntimeError(f"anthropic HTTP {e.code}: {detail}")
        except Exception as e:  # noqa
            raise RuntimeError(f"anthropic {type(e).__name__}: {e}")
        blocks = data.get("content") or []
        txt = "".join(b.get("text", "") for b in blocks if isinstance(b, dict) and b.get("type") == "text")
        if txt.strip():
            out.append(txt)
    return "\n\n".join(out)


def _no_reader_note(kind_label):
    """Actionable note when nothing could read a visual source."""
    if _VISION_PROVIDER == "claude" and _vision_on and not _anthropic_key():
        return (f"{kind_label} — set ANTHROPIC_API_KEY to read it with Claude vision, "
                "or run .claude/tools/setup-tools.sh to install tesseract as a fallback")
    return (f"{kind_label} — no extractable text layer (needs OCR; "
            "run .claude/tools/setup-tools.sh to install tesseract, then re-check)")


def _read_scanned_pdf(path):
    """Read an image-only/scanned PDF: Claude vision (if configured) else tesseract, cached.
    Returns (text, note, err)."""
    if not _OCR_ENABLED:
        return "", None, None
    tag = f"claude:{_VISION_MODEL}" if _vision_configured() else "tesseract"
    cache = _read_cache_path(path, tag)
    hit = _cache_read(cache)
    if hit is not None:
        return hit, ("read via Claude vision (was image-only scan; cached)" if tag.startswith("claude")
                     else "OCR'd (was image-only scan; cached)"), None
    if _ocr_deadline is not None and _time.monotonic() > _ocr_deadline:
        return "", None, ("image-only/scanned PDF — visual read deferred (pre-flight time budget); "
                          "it will be read automatically when the run starts")
    if not shutil.which("pdftoppm"):
        return "", None, None  # can't rasterise; caller shows the needs-OCR/setup-tools hint
    try:
        with tempfile.TemporaryDirectory() as td:
            pages = _render_pdf_pages(path, td)
            if not pages:
                return "", None, "image-only/scanned PDF — could not rasterise pages"
            if _vision_configured():
                try:
                    txt = _vision_transcribe(pages, os.path.basename(path))
                    if txt.strip():
                        _cache_write(cache, txt)
                        return txt, "read via Claude vision (was image-only scan)", None
                except Exception:
                    pass  # fall through to the tesseract floor
            if shutil.which("tesseract"):
                txt = _tesseract_pngs(pages)
                if txt.strip():
                    _cache_write(_read_cache_path(path, "tesseract"), txt)
                    return txt, "OCR'd (was image-only scan)", None
    except subprocess.TimeoutExpired:
        return "", None, f"image-only/scanned PDF — visual read timed out after {_OCR_PER_FILE_TIMEOUT}s"
    except Exception as e:  # noqa
        return "", None, f"image-only/scanned PDF — visual read failed ({type(e).__name__}: {e})"
    # a reader ran but found nothing (blank/graphic pages) vs no reader available at all
    if _vision_configured() or shutil.which("tesseract"):
        return "", None, "image-only/scanned PDF — no readable text found (blank or pure-graphic pages?)"
    return "", None, _no_reader_note("image-only/scanned PDF")


def _read_image_file(path):
    """Read a standalone image (screenshot / photographed table / chart): Claude vision (if configured)
    else tesseract, cached. Returns (text, note, err)."""
    if not _OCR_ENABLED:
        return "", None, None
    tag = f"claude:{_VISION_MODEL}" if _vision_configured() else "tesseract"
    cache = _read_cache_path(path, tag)
    hit = _cache_read(cache)
    if hit is not None:
        return hit, ("read via Claude vision (cached)" if tag.startswith("claude") else "OCR'd image (cached)"), None
    if _ocr_deadline is not None and _time.monotonic() > _ocr_deadline:
        return "", None, "image — visual read deferred (pre-flight time budget); done when the run starts"
    try:
        if _vision_configured():
            try:
                with tempfile.TemporaryDirectory() as td:
                    png = _normalize_image_to_png(path, td)
                    if png:
                        txt = _vision_transcribe([png], os.path.basename(path))
                        if txt.strip():
                            _cache_write(cache, txt)
                            return txt, "read via Claude vision", None
            except Exception:
                pass  # fall through to tesseract
        if shutil.which("tesseract"):
            r = subprocess.run(["tesseract", path, "stdout", "--psm", "1"],
                               capture_output=True, text=True, timeout=_OCR_PER_FILE_TIMEOUT)
            if r.returncode == 0 and r.stdout.strip():
                _cache_write(_read_cache_path(path, "tesseract"), r.stdout)
                return r.stdout, "OCR'd image", None
    except subprocess.TimeoutExpired:
        return "", None, f"image — visual read timed out after {_OCR_PER_FILE_TIMEOUT}s"
    except Exception as e:  # noqa
        return "", None, f"image — visual read failed ({type(e).__name__}: {e})"
    # a reader ran but found nothing (a photo/logo/graphic) vs no reader available at all
    if _vision_configured() or shutil.which("tesseract"):
        return "", None, "image — no readable text found (a photo/logo/graphic without machine text?)"
    return "", None, _no_reader_note("image")


def _read_pdf(path, max_pages=None):
    """Extract PDF text. Returns (text, err, note). Tries pdftotext -> pypdf; on a FULL read
    (max_pages is None — not the fast header sniff) that yields nothing, transcribes the scan
    (Claude vision or tesseract) once and caches it, so an image-only PDF becomes usable everywhere
    instead of a permanent 'needs OCR' fail. The header-sniff path (max_pages set) never does this."""
    txt, err = _read_pdf_text(path, max_pages)
    if txt.strip():
        return txt, None, None
    if max_pages is None:
        vtxt, note, verr = _read_scanned_pdf(path)
        if vtxt.strip():
            return vtxt, None, note
        if verr:
            err = verr
    return "", err, None


def _read_rtf(path):
    try:  # macOS textutil handles RTF (and HTML)
        r = subprocess.run(["textutil", "-convert", "txt", "-stdout", path],
                           capture_output=True, text=True, timeout=120)
        if r.returncode == 0:
            return r.stdout, None
        return "", f"textutil rc={r.returncode}"
    except FileNotFoundError:
        return "", "textutil not available"
    except Exception as e:  # noqa
        return "", f"{type(e).__name__}: {e}"


# ---------- binary-Word / MHTML / HTML (Capital IQ's mislabeled exports) ----------

def _html_to_text(b):
    """Strip an HTML/XHTML byte string to readable text (no external deps)."""
    t = b.decode("utf-8", "ignore") if isinstance(b, (bytes, bytearray)) else b
    t = re.sub(r"(?is)<(script|style|head)\b.*?</\1>", " ", t)
    t = re.sub(r"(?is)<(br|/p|/tr|/div|/h[1-6])\s*>", "\n", t)
    t = re.sub(r"(?is)</td>\s*<td[^>]*>", "\t", t)
    t = re.sub(r"(?s)<[^>]+>", " ", t)
    t = _htmlmod.unescape(t)
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n[ \t]*\n+", "\n\n", t)
    return t.strip()


def _read_doc(path):
    """Binary Word (OLE2 .doc) via macOS `textutil`. textutil keys off the file
    EXTENSION, so a CIQ transcript saved as `.rtf` but really OLE2 Word must be fed
    through a temp file with a `.doc` suffix. [fix F-EXTRACT — the CRM transcripts.]"""
    try:
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tf:
            tf.write(open(path, "rb").read())
            tmp = tf.name
        try:
            r = subprocess.run(["textutil", "-convert", "txt", "-stdout", tmp],
                               capture_output=True, text=True, timeout=180)
        finally:
            os.unlink(tmp)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout, None
        return "", f"textutil rc={r.returncode}"
    except FileNotFoundError:
        return "", "textutil not available (macOS-only); install antiword/libreoffice for .doc"
    except Exception as e:  # noqa
        return "", f"{type(e).__name__}: {e}"


def _read_mhtml(path):
    """MHTML — HTML wrapped in MIME, how Capital IQ exports filings (often a `.doc`
    extension). Parse the MIME, take the HTML part, strip to text.
    [fix F-EXTRACT — the CRM 10-K extracted as MHTML, 598K chars recovered.]"""
    try:
        msg = _email.message_from_bytes(open(path, "rb").read())
        htmlbytes = None
        for part in msg.walk():
            if part.get_content_type() in ("text/html", "application/xhtml+xml"):
                htmlbytes = part.get_payload(decode=True)
                break
        if htmlbytes is None:
            htmlbytes = msg.get_payload(decode=True) or open(path, "rb").read()
        txt = _html_to_text(htmlbytes)
        return (txt, None) if txt.strip() else ("", "mhtml: empty after html strip")
    except Exception as e:  # noqa
        return "", f"{type(e).__name__}: {e}"


def _read_html(path):
    try:
        return _html_to_text(open(path, "rb").read()), None
    except Exception as e:  # noqa
        return "", f"{type(e).__name__}: {e}"


# ---------- one-file inspect (UI / --list-json) ----------

def list_file(path):
    base = os.path.basename(path)
    ext = base.lower().rsplit(".", 1)[-1] if "." in base else ""
    methodology_reason = methodology_only_source_reason(
        base, sniff_text(path, METHODOLOGY_ONLY_SNIFF_CHARS),
    )
    if methodology_reason:
        return {"file": base, "ext": ext, "kind": "methodology-only", "status": "fail",
                "error": f"{METHODOLOGY_ONLY_ERROR}: {methodology_reason}", "sheets": []}
    fmt = sniff_format(path)  # content, not extension — CIQ mislabels workbooks
    if fmt in ("ole2", "zip"):
        wbext = "xls" if fmt == "ole2" else "xlsx"
        try:
            sheets = read_workbook(path, wbext)
            return {
                "file": base, "ext": ext, "kind": "workbook", "status": "ok",
                "sheets": [
                    {"name": nm, "rows": nr, "cols": nc, "cells": _nonempty_cells(rows)}
                    for (nm, nr, nc, rows) in sheets
                ],
            }
        except Exception as e:  # noqa — OLE2/zip Word doc, corrupt, or missing lib
            return {"file": base, "ext": ext, "kind": "document",
                    "status": "fail", "error": f"{type(e).__name__}: {e}", "sheets": []}
    return {"file": base, "ext": ext, "kind": fmt, "status": "skipped", "sheets": []}


def sniff_text(path, max_chars=16000):
    """Plain-text head of ANY supported file, for content-classification (the cockpit
    sniff). Workbook -> tab names + first rows; pdf -> first pages; rtf -> textutil;
    txt/csv/md -> head. One extractor, so the cockpit reads pdf/rtf the same way the
    pipeline does instead of guessing from raw bytes."""
    fmt = sniff_format(path)  # content, not extension
    try:
        if fmt in ("ole2", "zip"):
            wbext = "xls" if fmt == "ole2" else "xlsx"
            try:
                parts = []
                for (nm, _nr, _nc, rows) in read_workbook(path, wbext):
                    parts.append(nm)
                    for r in rows[:40]:
                        parts.append("\t".join(r))
                    if sum(len(x) for x in parts) > max_chars:
                        break
                return "\n".join(parts)[:max_chars]
            except Exception:  # noqa — OLE2/zip Word doc, not a sheet
                return (_read_doc(path)[0] or "")[:max_chars]
        if fmt == "pdf":
            return (_read_pdf(path, max_pages=12)[0] or "")[:max_chars]
        if fmt == "mime":
            return (_read_mhtml(path)[0] or "")[:max_chars]
        if fmt == "rtf":
            return (_read_rtf(path)[0] or "")[:max_chars]
        if fmt == "html":
            return (_read_html(path)[0] or "")[:max_chars]
        if fmt == "text":
            return open(path, errors="ignore").read(max_chars)
    except Exception:  # noqa — sniffing must never throw
        return ""
    return ""


# ---------- full pool extract ----------

_OUTPUT_LOCK_NAME = ".extract-pool.flock"
_GENERATION_DIR_NAME = ".extract-generations"
_GENERATION_BINDING_FORMAT = "python-json-sort-keys-compact-utf8/v1"
_LOCK_TIMEOUT_ENV = "EXTRACT_POOL_LOCK_TIMEOUT_S"
_DEFAULT_LOCK_TIMEOUT_S = 120.0
_SNAPSHOT_ATTEMPTS = 2

# A full-run supervisor can bind every later child to the exact pool generation
# admitted by the chain root.  All three variables are required together.  When
# present, ``extract_pool`` validates and reuses that immutable generation
# without enumerating or reading the live Drive pool.
FROZEN_POOL_DATA_PATH_ENV = "NOSTRA_FROZEN_POOL_DATA_PATH"
FROZEN_POOL_OUT_DIR_ENV = "NOSTRA_FROZEN_POOL_OUT_DIR"
FROZEN_POOL_GENERATION_ENV = "NOSTRA_FROZEN_POOL_GENERATION"


class PoolInventoryError(RuntimeError):
    """The pool could not be traversed completely, so an empty verdict is unsafe."""


class PoolChangedDuringExtraction(RuntimeError):
    """The complete pool changed between snapshot and publication."""


class FrozenPoolGenerationError(RuntimeError):
    """A supervisor-bound immutable pool generation is absent or does not verify."""


@contextmanager
def _pool_output_lease(out_dir, timeout_s=None, cancelled=None):
    """Serialize every publisher targeting one extraction directory.

    Full-run module children share ``<RUN_ROOT>/_pool_extracts``.  Without one
    retained kernel lease, two extractors can both miss the freshness cache and
    truncate the same manifest/extract files while the other is still reading
    or writing them.  The lock inode is persistent: a process exit releases the
    lease in the kernel, so no stale PID file can wedge later work.
    """
    os.makedirs(out_dir, exist_ok=True)
    directory = os.path.realpath(os.path.abspath(out_dir))
    lock_path = os.path.join(directory, _OUTPUT_LOCK_NAME)
    flags = os.O_RDWR | os.O_CREAT | getattr(os, "O_NOFOLLOW", 0)
    fd = None
    try:
        fd = os.open(lock_path, flags, 0o600)
        opened = os.fstat(fd)
        named = os.lstat(lock_path)
        if (not stat.S_ISREG(opened.st_mode)
                or stat.S_ISLNK(named.st_mode)
                or opened.st_uid != os.getuid()
                or opened.st_nlink != 1
                or stat.S_IMODE(opened.st_mode) & 0o077
                or (opened.st_dev, opened.st_ino) != (named.st_dev, named.st_ino)):
            raise RuntimeError(f"unsafe extraction lock at {lock_path}")
        if timeout_s is None:
            raw_timeout = os.environ.get(_LOCK_TIMEOUT_ENV, str(_DEFAULT_LOCK_TIMEOUT_S))
            try:
                timeout_s = float(raw_timeout)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{_LOCK_TIMEOUT_ENV} must be a finite non-negative number") from exc
        if not isinstance(timeout_s, (int, float)) or timeout_s < 0 or timeout_s != timeout_s:
            raise ValueError("extraction lock timeout must be a finite non-negative number")
        # Never inherit an unbounded wait from a caller or environment typo.  A
        # crashed holder releases flock in the kernel; a live but wedged holder
        # must surface as a technical failure instead of freezing every module.
        timeout_s = min(float(timeout_s), 300.0)
        deadline = _time.monotonic() + timeout_s
        while True:
            if cancelled is not None and cancelled():
                raise InterruptedError("extraction lock wait cancelled")
            try:
                fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                break
            except OSError as exc:
                if exc.errno not in (errno.EACCES, errno.EAGAIN):
                    raise
                remaining = deadline - _time.monotonic()
                if remaining <= 0:
                    raise TimeoutError(
                        f"timed out after {timeout_s:g}s waiting for extraction lock at {lock_path}"
                    ) from None
                _time.sleep(min(0.05, remaining))
        yield
    finally:
        if fd is not None:
            os.close(fd)

# v3: _file_guard_identity dropped ctime_ns (a Drive xattr touch on one file was discarding the
# whole cached manifest). The shape changed, so every v2 guard must be treated as stale rather than
# compared field-by-field against a differently-shaped dict.
_CACHE_GUARD_VERSION = 3

def _complete_pool_inventory(data_path):
    """Return every safe pool file, or fail if traversal was incomplete.

    An I/O error is not an empty directory.  The old generator swallowed root,
    walk, directory-lstat, and file-lstat failures, allowing a transient Drive
    problem to collapse a populated pool into ``zero_files``.  This routine
    completes traversal before returning anything and surfaces every uncertain
    branch as a technical error.
    """
    # Resolve the configured root once (the root itself may intentionally be a
    # Drive symlink), then never recurse through a descendant symlink. Yield
    # the caller's lexical root so existing relpath/provenance labels remain
    # stable. Every consumer that reads bytes re-opens through pinned dirfds.
    lexical_root = os.path.abspath(data_path)
    real_root = os.path.realpath(lexical_root)
    try:
        root_info = os.lstat(real_root)
    except OSError as exc:
        raise PoolInventoryError(f"cannot inspect pool root {real_root}: {exc}") from exc
    if stat.S_ISLNK(root_info.st_mode) or not stat.S_ISDIR(root_info.st_mode):
        raise PoolInventoryError(f"pool root is not a directory: {real_root}")

    def _raise_walk_error(exc):
        raise exc

    inventory = []
    try:
        walker = os.walk(real_root, topdown=True, followlinks=False, onerror=_raise_walk_error)
        for directory, dirs, files in walker:
            safe_dirs = []
            for name in sorted(dirs):
                candidate = os.path.join(directory, name)
                try:
                    info = os.lstat(candidate)
                except OSError as exc:
                    raise PoolInventoryError(f"cannot inspect pool directory {candidate}: {exc}") from exc
                if stat.S_ISDIR(info.st_mode) and not stat.S_ISLNK(info.st_mode):
                    safe_dirs.append(name)
            dirs[:] = safe_dirs
            for name in sorted(files):
                real_path = os.path.join(directory, name)
                try:
                    info = os.lstat(real_path)
                except OSError as exc:
                    raise PoolInventoryError(f"cannot inspect pool file {real_path}: {exc}") from exc
                if (stat.S_ISLNK(info.st_mode) or not stat.S_ISREG(info.st_mode)
                        or info.st_nlink != 1):
                    continue
                rel = os.path.relpath(real_path, real_root)
                if any(part in {"", ".", ".."} for part in rel.split(os.sep)):
                    raise PoolInventoryError(f"unsafe relative path discovered in pool: {rel!r}")
                p = os.path.join(lexical_root, rel)
                # Skip engine-written output folders (the memos/thesis/dossier saved back into the
                # company's Drive folder), marked by a .nostradamus_output sentinel.  A marker lookup
                # that fails for any reason other than absence makes the inventory uncertain.
                marker = os.path.join(os.path.dirname(p), ".nostradamus_output")
                try:
                    os.lstat(marker)
                except FileNotFoundError:
                    pass
                except OSError as exc:
                    raise PoolInventoryError(f"cannot inspect output marker {marker}: {exc}") from exc
                else:
                    continue
                inventory.append(p)
    except PoolInventoryError:
        raise
    except OSError as exc:
        raise PoolInventoryError(f"pool traversal failed under {real_root}: {exc}") from exc
    return tuple(inventory)


def iter_pool_files(data_path):
    """Iterate only after a complete, error-aware traversal has succeeded."""
    yield from _complete_pool_inventory(data_path)


def _pool_rel(data_path, path):
    rel = os.path.relpath(path, os.path.abspath(data_path))
    if any(part in {"", ".", ".."} for part in rel.split(os.sep)):
        raise PoolInventoryError(f"unsafe pool-relative path: {rel!r}")
    return rel


def _capture_pool_snapshot(data_path, inventory, snapshot_pool):
    """Copy the already-complete inventory once and bind every byte by SHA-256."""
    captured = []
    for path in inventory:
        rel = _pool_rel(data_path, path)
        try:
            snapshot_path = _snapshot_pool_file(data_path, rel, snapshot_pool)
            with open(snapshot_path, "rb") as handle:
                raw = handle.read()
            live = os.stat(path, follow_symlinks=False)
        except (OSError, ValueError) as exc:
            raise PoolChangedDuringExtraction(
                f"pool changed or became unreadable while snapshotting {rel}: {exc}"
            ) from exc
        captured.append({
            "path": path,
            "rel": rel.replace(os.sep, "/"),
            "snapshot_path": snapshot_path,
            "sha256": hashlib.sha256(raw).hexdigest(),
            "size": len(raw),
            "mtime_ns": live.st_mtime_ns,
        })
    return tuple(captured)


def _revalidate_pool_inventory(data_path, captured):
    """Prove the complete live inventory still equals the captured transaction."""
    current_paths = _complete_pool_inventory(data_path)
    expected = [entry["rel"] for entry in captured]
    current_rels = [_pool_rel(data_path, path).replace(os.sep, "/") for path in current_paths]
    if current_rels != expected:
        raise PoolChangedDuringExtraction(
            "pool file set changed during extraction "
            f"(captured={expected!r}, current={current_rels!r})"
        )

    verified = []
    for path, captured_entry in zip(current_paths, captured):
        rel = captured_entry["rel"]
        try:
            raw = _read_pool_regular_bytes(data_path, rel)
            live = os.stat(path, follow_symlinks=False)
        except (OSError, ValueError) as exc:
            raise PoolChangedDuringExtraction(
                f"pool changed or became unreadable before publication at {rel}: {exc}"
            ) from exc
        digest = hashlib.sha256(raw).hexdigest()
        if len(raw) != captured_entry["size"] or digest != captured_entry["sha256"]:
            raise PoolChangedDuringExtraction(f"pool bytes changed during extraction at {rel}")
        verified.append({
            **captured_entry,
            "path": path,
            "size": len(raw),
            "mtime_ns": live.st_mtime_ns,
        })
    return tuple(verified)


def _generation_reference(generation_digest, local_name):
    local = str(local_name).replace(os.sep, "/").lstrip("/")
    if any(part in {"", ".", ".."} for part in local.split("/")):
        raise ValueError(f"unsafe generation artifact name: {local_name!r}")
    return f"{_GENERATION_DIR_NAME}/{generation_digest}/{local}"


def _localize_generation_sources(sources):
    """Copy sources and strip a generation prefix for digest calculation."""
    cloned = json.loads(json.dumps(sources, default=str))
    for source in cloned:
        names = [(source, "extract")]
        names.extend((sheet, "extract") for sheet in (source.get("sheets") or []) if isinstance(sheet, dict))
        for owner, key in names:
            value = owner.get(key)
            if not isinstance(value, str) or value == "(original)":
                continue
            parts = value.replace(os.sep, "/").split("/")
            if len(parts) >= 3 and parts[0] == _GENERATION_DIR_NAME:
                owner[key] = "/".join(parts[2:])
    return cloned


def _localize_original_sources(sources, raw_prefix):
    """Give in-place text an immutable local artifact instead of a Drive sentinel."""
    cloned = json.loads(json.dumps(sources, default=str))
    for source in cloned:
        if not isinstance(source, dict) or source.get("extract") != "(original)":
            continue
        rel = str(source.get("path") or source.get("file") or "").replace(os.sep, "/")
        if not rel or any(part in {"", ".", ".."} for part in rel.split("/")):
            raise ValueError(f"unsafe in-place source path: {rel!r}")
        source["extract"] = f"{raw_prefix.rstrip('/')}/{rel}"
    return cloned


def _prefix_generation_sources(sources, generation_digest):
    cloned = json.loads(json.dumps(sources, default=str))
    for source in cloned:
        names = [(source, "extract")]
        names.extend((sheet, "extract") for sheet in (source.get("sheets") or []) if isinstance(sheet, dict))
        for owner, key in names:
            value = owner.get(key)
            if isinstance(value, str):
                owner[key] = _generation_reference(generation_digest, value)
    return cloned


def _hash_generation_artifacts(directory):
    artifacts = {}
    for root, dirs, files in os.walk(directory, topdown=True, followlinks=False):
        dirs[:] = sorted(d for d in dirs if not os.path.islink(os.path.join(root, d)))
        for name in sorted(files):
            path = os.path.join(root, name)
            info = os.lstat(path)
            if stat.S_ISLNK(info.st_mode) or not stat.S_ISREG(info.st_mode):
                raise ValueError(f"unsafe staged generation artifact: {path}")
            rel = os.path.relpath(path, directory).replace(os.sep, "/")
            with open(path, "rb") as handle:
                artifacts[rel] = hashlib.sha256(handle.read()).hexdigest()
    return artifacts


def _generation_binding_payload(
        data_path, out_dir, captured, sources, entities, artifacts, vision_mode,
        offline_extraction_complete, raw_prefix):
    return {
        "schema_version": "pool-generation/v2",
        # This deliberately changes the content address from the first v2
        # receipts, which did not retain the bytes Python hashed.  A stale v2
        # cache therefore rebuilds into a new directory instead of colliding
        # with an unverifiable legacy directory at the old digest.
        "binding_format": _GENERATION_BINDING_FORMAT,
        "data_path": os.path.abspath(data_path),
        "out_dir": os.path.abspath(out_dir),
        "vision_mode": bool(vision_mode),
        "offline_extraction_complete": bool(offline_extraction_complete),
        "raw_prefix": raw_prefix,
        "inputs": {entry["rel"]: entry["sha256"] for entry in captured},
        "sources": _localize_generation_sources(sources),
        "entities": entities,
        "artifacts": dict(sorted(artifacts.items())),
    }


def _generation_binding_json(payload):
    """Return the exact Python JSON text used as the generation content address.

    Keep these bytes in the receipt.  Another runtime can hash and parse them,
    but must not try to reproduce Python's number spelling (for example 5.0 or
    1e-07) with its own JSON serializer.
    """
    return json.dumps(
        payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False,
        allow_nan=False,
    )


def _generation_digest_from_json(binding_json):
    if not isinstance(binding_json, str):
        raise ValueError("generation binding JSON must be text")
    return hashlib.sha256(binding_json.encode("utf-8")).hexdigest()


def _generation_digest(payload):
    return _generation_digest_from_json(_generation_binding_json(payload))


def _frozen_pool_binding_from_env():
    values = {
        "data_path": os.environ.get(FROZEN_POOL_DATA_PATH_ENV),
        "out_dir": os.environ.get(FROZEN_POOL_OUT_DIR_ENV),
        "generation": os.environ.get(FROZEN_POOL_GENERATION_ENV),
    }
    present = [value is not None for value in values.values()]
    if not any(present):
        return None
    if not all(present):
        raise FrozenPoolGenerationError(
            "frozen pool reuse requires all of "
            f"{FROZEN_POOL_DATA_PATH_ENV}, {FROZEN_POOL_OUT_DIR_ENV}, "
            f"and {FROZEN_POOL_GENERATION_ENV}"
        )
    if not re.fullmatch(r"[0-9a-f]{64}", values["generation"] or ""):
        raise FrozenPoolGenerationError("frozen pool generation must be a lowercase SHA-256 digest")
    return values


def _assert_plain_directory(path, label, *, read_only=False):
    try:
        info = os.lstat(path)
    except OSError as exc:
        raise FrozenPoolGenerationError(f"{label} is unavailable") from exc
    if stat.S_ISLNK(info.st_mode) or not stat.S_ISDIR(info.st_mode):
        raise FrozenPoolGenerationError(f"{label} is not a plain directory")
    if read_only and os.name != "nt" and info.st_mode & 0o222:
        raise FrozenPoolGenerationError(f"{label} is writable")


def _plain_generation_artifact(generation_dir, rel, *, require_read_only=True):
    """Resolve one manifest member without following any symlink component."""
    if (not isinstance(rel, str) or rel.startswith("/")
            or any(part in {"", ".", ".."} for part in rel.split("/"))):
        raise FrozenPoolGenerationError("frozen pool artifact path is malformed")
    current = generation_dir
    for part in rel.split("/")[:-1]:
        current = os.path.join(current, part)
        _assert_plain_directory(
            current, f"frozen pool artifact directory {rel}",
            read_only=require_read_only,
        )
    path = os.path.join(generation_dir, *rel.split("/"))
    try:
        info = os.lstat(path)
    except OSError as exc:
        raise FrozenPoolGenerationError(f"frozen pool artifact is missing: {rel}") from exc
    if stat.S_ISLNK(info.st_mode) or not stat.S_ISREG(info.st_mode):
        raise FrozenPoolGenerationError(f"frozen pool artifact is not a plain file: {rel}")
    if require_read_only and os.name != "nt" and info.st_mode & 0o222:
        raise FrozenPoolGenerationError(f"frozen pool artifact is writable: {rel}")
    return path


def _read_plain_generation_artifact(generation_dir, rel, *, require_read_only=True):
    """Read one plain generation member from a pinned file descriptor.

    ``_plain_generation_artifact`` rejects an existing symlink, but a path can
    still be replaced between that lstat and a later ``open``.  Open the final
    component with ``O_NOFOLLOW`` where the platform supports it, validate the
    descriptor itself, and read from that descriptor.  Callers still compare
    the returned bytes with the generation's digest binding.
    """
    path = _plain_generation_artifact(
        generation_dir, rel, require_read_only=require_read_only,
    )
    flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
    try:
        fd = os.open(path, flags)
    except OSError as exc:
        raise FrozenPoolGenerationError(
            f"frozen pool artifact is unavailable: {rel}"
        ) from exc
    try:
        info = os.fstat(fd)
        if not stat.S_ISREG(info.st_mode):
            raise FrozenPoolGenerationError(
                f"frozen pool artifact is not a plain file: {rel}"
            )
        if require_read_only and os.name != "nt" and info.st_mode & 0o222:
            raise FrozenPoolGenerationError(
                f"frozen pool artifact is writable: {rel}"
            )
        chunks = []
        while True:
            chunk = os.read(fd, 1024 * 1024)
            if not chunk:
                break
            chunks.append(chunk)
        return b"".join(chunks)
    finally:
        os.close(fd)


def _seal_generation_tree(generation_dir):
    """Make one published content-addressed generation OS read-only."""
    if os.name == "nt":
        return
    for root, dirs, files in os.walk(generation_dir, topdown=False, followlinks=False):
        for name in files:
            path = os.path.join(root, name)
            info = os.lstat(path)
            if stat.S_ISLNK(info.st_mode) or not stat.S_ISREG(info.st_mode):
                raise ValueError(f"unsafe generation member while sealing: {path}")
            os.chmod(path, 0o444, follow_symlinks=False)
        for name in dirs:
            path = os.path.join(root, name)
            info = os.lstat(path)
            if stat.S_ISLNK(info.st_mode) or not stat.S_ISDIR(info.st_mode):
                raise ValueError(f"unsafe generation directory while sealing: {path}")
            os.chmod(path, 0o555, follow_symlinks=False)
        os.chmod(root, 0o555, follow_symlinks=False)


def _verified_generation_manifest(
        data_path, out_dir, digest, *, require_read_only=True, require_complete=False):
    """Load and verify one immutable generation without touching the live pool."""
    generation_parent = os.path.join(os.path.abspath(out_dir), _GENERATION_DIR_NAME)
    generation_dir = os.path.join(generation_parent, digest)
    _assert_plain_directory(generation_parent, "frozen pool generation parent")
    _assert_plain_directory(
        generation_dir, "frozen pool generation", read_only=require_read_only,
    )
    try:
        manifest = json.loads(_read_plain_generation_artifact(
            generation_dir, "manifest.json", require_read_only=require_read_only,
        ).decode("utf-8"))
    except Exception as exc:
        raise FrozenPoolGenerationError(
            f"frozen pool generation {digest} is unavailable or unreadable"
        ) from exc
    generation = manifest.get("generation") if isinstance(manifest, dict) else None
    if not isinstance(generation, dict) or generation.get("digest") != digest:
        raise FrozenPoolGenerationError("frozen pool manifest generation does not match its binding")
    if generation.get("schema_version") != "pool-generation/v2":
        raise FrozenPoolGenerationError("frozen pool generation uses an unsupported legacy schema")
    if require_complete and manifest.get("offline_extraction_complete") is not True:
        raise FrozenPoolGenerationError("frozen pool generation contains a time-budgeted partial extraction")
    if os.path.abspath(str(manifest.get("data_path", ""))) != os.path.abspath(data_path):
        raise FrozenPoolGenerationError("frozen pool manifest is bound to a different data path")
    if os.path.abspath(str(manifest.get("out_dir", ""))) != os.path.abspath(out_dir):
        raise FrozenPoolGenerationError("frozen pool manifest is bound to a different output path")
    artifacts = generation.get("artifacts")
    inputs = generation.get("inputs")
    raw_prefix = generation.get("raw_prefix")
    binding_json = generation.get("binding_json")
    if (not isinstance(artifacts, dict) or not isinstance(inputs, dict)
            or not isinstance(raw_prefix, str) or not isinstance(binding_json, str)
            or any(part in {"", ".", ".."} for part in raw_prefix.split("/"))):
        raise FrozenPoolGenerationError("frozen pool manifest lacks its immutable binding")
    for rel, expected in artifacts.items():
        if (not isinstance(rel, str) or not isinstance(expected, str)
                or not re.fullmatch(r"[0-9a-f]{64}", expected)):
            raise FrozenPoolGenerationError("frozen pool artifact binding is malformed")
        content = _read_plain_generation_artifact(
            generation_dir, rel, require_read_only=require_read_only,
        )
        actual = hashlib.sha256(content).hexdigest()
        if actual != expected:
            raise FrozenPoolGenerationError(f"frozen pool artifact changed: {rel}")
    for rel, expected in inputs.items():
        if (not isinstance(rel, str) or not isinstance(expected, str)
                or not re.fullmatch(r"[0-9a-f]{64}", expected)):
            raise FrozenPoolGenerationError("frozen pool input binding is malformed")
        raw_rel = f"{raw_prefix}/{rel}"
        if artifacts.get(raw_rel) != expected:
            raise FrozenPoolGenerationError(f"frozen pool raw input is not exactly bound: {rel}")
    exact_prefix = f"{_GENERATION_DIR_NAME}/{digest}/"
    for source in manifest.get("sources") or []:
        if not isinstance(source, dict):
            raise FrozenPoolGenerationError("frozen pool source binding is malformed")
        references = [source.get("extract")]
        references.extend(
            sheet.get("extract") for sheet in (source.get("sheets") or [])
            if isinstance(sheet, dict)
        )
        for reference in references:
            if reference is None:
                continue
            if not isinstance(reference, str) or not reference.startswith(exact_prefix):
                raise FrozenPoolGenerationError("frozen pool extract escapes its bound generation")
            local = reference[len(exact_prefix):]
            if artifacts.get(local) is None:
                raise FrozenPoolGenerationError("frozen pool extract is not a bound artifact")
    payload = {
        "schema_version": generation.get("schema_version"),
        "binding_format": _GENERATION_BINDING_FORMAT,
        "data_path": os.path.abspath(data_path),
        "out_dir": os.path.abspath(out_dir),
        "vision_mode": bool(manifest.get("vision_mode")),
        "offline_extraction_complete": bool(manifest.get("offline_extraction_complete")),
        "raw_prefix": raw_prefix,
        "inputs": inputs,
        "sources": _localize_generation_sources(manifest.get("sources") or []),
        "entities": manifest.get("entities") or [],
        "artifacts": dict(sorted(artifacts.items())),
    }
    try:
        parsed_binding = json.loads(binding_json)
    except (TypeError, ValueError) as exc:
        raise FrozenPoolGenerationError(
            "frozen pool generation binding JSON is unreadable"
        ) from exc
    # The digest authenticates the exact retained UTF-8 bytes.  The two
    # comparisons then prove those bytes are Python's canonical spelling of
    # the independently reconstructed manifest/artifact/input/source/entity
    # binding.  In particular, int/float spelling cannot disappear behind
    # Python's loose ``5 == 5.0`` equality.
    if _generation_digest_from_json(binding_json) != digest:
        raise FrozenPoolGenerationError("frozen pool generation digest does not verify")
    if parsed_binding != payload or binding_json != _generation_binding_json(payload):
        raise FrozenPoolGenerationError("frozen pool generation binding does not match its manifest")
    return manifest


def _read_bound_generation_artifact(out_dir, manifest, rel):
    """Re-read and verify a generation member immediately before projection."""
    generation = manifest.get("generation") if isinstance(manifest, dict) else None
    digest = generation.get("digest") if isinstance(generation, dict) else None
    artifacts = generation.get("artifacts") if isinstance(generation, dict) else None
    expected = artifacts.get(rel) if isinstance(artifacts, dict) else None
    if (not isinstance(digest, str) or not isinstance(expected, str)
            or not re.fullmatch(r"[0-9a-f]{64}", expected)):
        raise FrozenPoolGenerationError(
            f"frozen pool artifact is not bound by the generation: {rel}"
        )
    generation_dir = os.path.join(out_dir, _GENERATION_DIR_NAME, digest)
    content = _read_plain_generation_artifact(generation_dir, rel)
    if hashlib.sha256(content).hexdigest() != expected:
        raise FrozenPoolGenerationError(f"frozen pool artifact changed: {rel}")
    return content


def _reuse_frozen_generation(data_path, out_dir, digest, corpus_path=None):
    manifest = _verified_generation_manifest(
        data_path, out_dir, digest, require_complete=True,
    )
    if corpus_path:
        _atomic_bytes(
            corpus_path,
            _read_bound_generation_artifact(out_dir, manifest, "corpus.txt"),
        )
    print(f"[extract_pool] frozen generation {digest} reused; live pool was not read")
    return manifest


def _file_guard_identity(path, *, content_hash=False):
    """Stable local identity; hash only derived/cache text, never every raw filing.

    ctime_ns is deliberately absent. This dict is a cache-FRESHNESS signal, compared whole by
    is_fresh() across every pool file; the pool is a Google Drive mirror, so a single xattr touch on
    a single file used to move ctime_ns and discard the entire cached manifest. Measured on the live
    INDIAMART pool: exactly one of 85 entries differed, in ctime_ns alone, with size and mtime_ns
    byte-identical — turning a ~1.3s cache hit into a ~24s full re-extraction, on essentially every
    run. size + mtime_ns is the standard freshness pair and is what actually tracks content; the
    integrity story for pool bytes is carried by the pinned reads and the attested-digest binding at
    extraction time, not by this heuristic.
    """
    st = os.stat(path)
    identity = {
        "size": st.st_size,
        "mtime_ns": st.st_mtime_ns,
    }
    if content_hash:
        digest = hashlib.sha256()
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                digest.update(chunk)
        identity["sha256"] = digest.hexdigest()
    return identity


def _visual_cache_path(path, tag):
    """Pure counterpart of `_read_cache_path` (does not create the cache dir)."""
    ident = _visual_source_identity(path)
    return os.path.join(
        _OCR_CACHE_DIR, hashlib.sha1(f"{tag}|{ident}".encode()).hexdigest() + ".txt",
    )


def _tool_guard_identity(name):
    path = shutil.which(name)
    if not path:
        return None
    try:
        return {"path": os.path.realpath(path), **_file_guard_identity(path)}
    except OSError:
        return {"path": os.path.realpath(path), "unreadable": True}


def _cache_guard_snapshot(data_path, out_dir, manifest, *, captured=None, extract_root=None):
    """Identity inputs which can change a cached manifest's evidence verdict.

    Raw pool inputs use size + nanosecond mtime/ctime so an ordinary edit is
    cheap to detect.  Derived extracts and OCR/vision caches are hashed because
    downstream readers must never trust stale/tampered extracted text merely
    because `manifest.json` is newer.  Reader availability is also part of the
    identity: installing OCR can make a previously unreadable renamed scan
    newly identifiable as methodology-only and therefore forces one rebuild.
    """
    pool_inputs = {}
    visual_caches = {}
    inventory_rows = captured
    if inventory_rows is None:
        inventory_rows = tuple({
            "path": path,
            "rel": os.path.relpath(path, data_path).replace(os.sep, "/"),
            "snapshot_path": path,
            **_file_guard_identity(path),
        } for path in _complete_pool_inventory(data_path))
    for entry in inventory_rows:
        path = entry.get("snapshot_path") or entry["path"]
        rel = entry["rel"]
        pool_inputs[rel] = {
            "size": entry["size"],
            "mtime_ns": entry["mtime_ns"],
        }
        if _is_sidecar(os.path.basename(path)):
            continue
        ext = path.lower().rsplit(".", 1)[-1] if "." in os.path.basename(path) else ""
        fmt = sniff_format(path)
        if fmt not in {"pdf", "image"} and ext not in IMAGE_EXTS:
            continue
        for tag in ("tesseract", f"claude:{_VISION_MODEL}"):
            cache = _visual_cache_path(path, tag)
            if os.path.isfile(cache):
                visual_caches[f"{rel}|{tag}"] = _file_guard_identity(cache, content_hash=True)

    approved_extracts = {}
    for source in (manifest or {}).get("sources", []):
        names = []
        extract = source.get("extract") if isinstance(source, dict) else None
        if isinstance(extract, str) and extract.endswith(".txt"):
            names.append(extract)
        for sheet in ((source.get("sheets") or []) if isinstance(source, dict) else []):
            sheet_extract = sheet.get("extract") if isinstance(sheet, dict) else None
            if isinstance(sheet_extract, str) and sheet_extract.endswith(".txt"):
                names.append(sheet_extract)
        for name in names:
            if extract_root is not None:
                local = name.replace(os.sep, "/").split("/")
                if len(local) >= 3 and local[0] == _GENERATION_DIR_NAME:
                    local = local[2:]
                path = os.path.join(extract_root, *local)
            else:
                path = os.path.join(out_dir, name)
            if not os.path.isfile(path):
                approved_extracts[name] = {"missing": True}
            else:
                approved_extracts[name] = _file_guard_identity(path, content_hash=True)
    # Extraction imports its methodology-only identity rules, manifest validation, eligibility clock and
    # projection verification from these shared modules at run time, so they are as much a part of the
    # verdict as the pool bytes are. is_fresh only mtime-checks extract_pool.py itself, so without hashing
    # them a TIGHTENED policy left every cached manifest valid: evidence that has just become forbidden
    # keeps being served as usable until something unrelated happens to touch the pool.
    policy_modules = {}
    for policy_path in (
        os.path.join(SCRIPTS_DIR, "connector_contract.py"),
        os.path.join(SCRIPTS_DIR, "connector_vintages.py"),
        os.path.join(TOOLS_DIR, "run_connectors.py"),
    ):
        rel = os.path.relpath(policy_path, REPO_ROOT).replace(os.sep, "/")
        policy_modules[rel] = (
            _file_guard_identity(policy_path, content_hash=True)
            if os.path.isfile(policy_path) else {"missing": True}
        )
    return {
        "version": _CACHE_GUARD_VERSION,
        "pool_inputs": pool_inputs,
        "approved_extracts": approved_extracts,
        "visual_caches": visual_caches,
        "policy_modules": policy_modules,
        "reader_state": {
            "ocr_enabled": bool(_OCR_ENABLED),
            "tesseract": _tool_guard_identity("tesseract"),
            "pdftoppm": _tool_guard_identity("pdftoppm"),
            "vision_configured": bool(_vision_configured()),
            "vision_model": _VISION_MODEL if _vision_configured() else None,
        },
    }


def _cached_canonical_integrity(data_path, manifest, decision_at):
    """Revalidate v2 projections against live heads, code, and release clocks."""
    try:
        if SCRIPTS_DIR not in sys.path:
            sys.path.insert(0, SCRIPTS_DIR)
        from connector_contract import canonical_json_bytes

        provenance = _collect_sidecars(data_path, decision_at)
        for source in (manifest or {}).get("sources", []):
            if not isinstance(source, dict):
                return False
            rel = str(source.get("path") or source.get("file") or "").replace("/", os.sep)
            if not rel:
                continue
            if not _is_external_rel(rel):
                continue
            owner = _v2_projection_manifest(data_path, rel)
            prov = provenance.get(rel)
            if owner is None:
                # A cached v2 claim whose manifest is no longer valid must not
                # degrade into an ordinary path-derived external document.
                if isinstance(prov, dict) and prov.get("_integrity_error"):
                    if not (source.get("status") == "fail"
                            and source.get("provenance") == {"integrity_status": "failed", "usable": False}
                            and prov["_integrity_error"] in str(source.get("error") or "")):
                        return False
                continue
            if not isinstance(prov, dict) or prov.get("_integrity_error"):
                if (isinstance(prov, dict) and prov.get("_integrity_error")
                        and source.get("status") == "fail"
                        and source.get("provenance") == {"integrity_status": "failed", "usable": False}
                        and prov["_integrity_error"] in str(source.get("error") or "")):
                    continue
                return False
            if prov.get("_projection_state") == "superseded":
                if (source.get("status") != "skipped"
                        or source.get("provenance") != {
                            "integrity_status": "superseded", "usable": False,
                        }):
                    return False
                continue
            if prov.get("_projection_state") == "disputed":
                if (source.get("status") != "skipped"
                        or source.get("provenance") != {
                            "integrity_status": "disputed", "usable": False,
                        }):
                    return False
                continue
            if prov.get("_projection_state") == "expired":
                if (source.get("status") != "skipped"
                        or source.get("provenance") != {
                            "eligibility_status": "expired", "usable": False,
                        }):
                    return False
                continue
            current_provenance = _enforce_tier_ceiling(
                {key: value for key, value in prov.items() if not key.startswith("_")}
            )
            if canonical_json_bytes(source.get("provenance")) != canonical_json_bytes(current_provenance):
                return False
        return True
    except Exception:
        return False


def is_fresh(out_dir, data_path, script_path, now=None):
    decision_at = _decision_time_utc(now)
    man = os.path.join(out_dir, "manifest.json")
    if not os.path.exists(man):
        return None
    man_m = os.path.getmtime(man)
    newest = os.path.getmtime(script_path)
    for p in iter_pool_files(data_path):
        newest = max(newest, os.path.getmtime(p))
    if man_m >= newest:
        try:
            cached = json.load(open(man))
        except Exception:  # noqa
            return None
        generation = cached.get("generation") if isinstance(cached, dict) else None
        digest = generation.get("digest") if isinstance(generation, dict) else None
        # A legacy flat cache is not an admission receipt.  Rebuild it into an
        # immutable raw+derived generation before any caller can treat it as
        # fresh, even when its old timestamps and guard still happen to match.
        if (not isinstance(generation, dict)
                or generation.get("schema_version") != "pool-generation/v2"
                or not isinstance(generation.get("raw_prefix"), str)
                or not isinstance(digest, str)
                or not re.fullmatch(r"[0-9a-f]{64}", digest)
                or not os.path.isfile(os.path.join(
                    out_dir, _GENERATION_DIR_NAME, digest, "manifest.json",
                ))):
            return None
        pointer_guard = cached.get("cache_guard") if isinstance(cached, dict) else None
        try:
            # Root manifest.json is only a replaceable pointer.  Statuses,
            # entities, provenance and source references used by readiness must
            # come from the self-verifying immutable generation, never from a
            # root-only edit that happens to preserve timestamps/cache shape.
            cached = _verified_generation_manifest(data_path, out_dir, digest)
        except FrozenPoolGenerationError:
            return None
        if (not isinstance(pointer_guard, dict)
                or pointer_guard.get("version") != _CACHE_GUARD_VERSION):
            return None
        try:
            # The mutable root pointer may carry current mtime freshness (a
            # deterministic rebuild can reuse the same immutable digest after
            # Drive rewrites identical bytes). Recompute it from trusted
            # generation references, but never consume root statuses/entities.
            if pointer_guard != _cache_guard_snapshot(data_path, out_dir, cached):
                return None
        except Exception:
            return None
        if not _cached_canonical_integrity(data_path, cached, decision_at):
            return None
        return cached
    return None


def extract_pool(
        data_path, out_dir, force=False, corpus_path=None, vision=None, now=None,
        ocr_budget_s=None):
    """Build or reuse one pool extraction under its output-directory lease."""
    frozen = _frozen_pool_binding_from_env()
    if frozen is not None:
        if os.path.abspath(data_path) != os.path.abspath(frozen["data_path"]):
            raise FrozenPoolGenerationError("extractor data path differs from supervisor frozen binding")
        if os.path.abspath(out_dir) != os.path.abspath(frozen["out_dir"]):
            raise FrozenPoolGenerationError("extractor output path differs from supervisor frozen binding")
        if force:
            raise FrozenPoolGenerationError("a frozen pool generation cannot be force-rebuilt")
        return _reuse_frozen_generation(
            data_path, out_dir, frozen["generation"], corpus_path=corpus_path,
        )
    with _pool_output_lease(out_dir):
        for attempt in range(_SNAPSHOT_ATTEMPTS):
            try:
                return _extract_pool_locked(
                    data_path, out_dir, force=force, corpus_path=corpus_path,
                    vision=vision, now=now, ocr_budget_s=ocr_budget_s,
                )
            except PoolChangedDuringExtraction:
                # Python 3.14 can retain the exception frame long enough that
                # TemporaryDirectory finalizers have not run before the next
                # bounded snapshot attempt.  Remove only this extractor's
                # private direct-child build directories while the output
                # lease is still held.
                for pending in glob.glob(os.path.join(out_dir, ".extract-build-*")):
                    shutil.rmtree(pending)
                if attempt + 1 >= _SNAPSHOT_ATTEMPTS:
                    raise
                print("[extract_pool] pool changed during snapshot; retrying one complete transaction")


def _extract_pool_locked(
        data_path, out_dir, force=False, corpus_path=None, vision=None, now=None,
        ocr_budget_s=None):
    script_path = os.path.abspath(__file__)
    decision_at = _decision_time_utc(now)
    # Vision (Claude) reads image-only PDFs + images. OFF in the readiness pre-flight (vision=False,
    # no token spend before the user proceeds); ON in-run (vision=None -> env default). tesseract stays
    # the offline floor either way.
    global _ocr_deadline, _vision_on
    _vision_on = _VISION_ON_ENV if vision is None else bool(vision)
    # A caller can explicitly request 0 = unbounded.  Readiness does so because
    # its generation becomes the immutable input for every later child; it may
    # never freeze a "deferred until the run" OCR row.  Other callers retain the
    # configured pool budget when this argument is omitted.
    effective_ocr_budget_s = (
        _OCR_POOL_BUDGET_S if ocr_budget_s is None else max(0, float(ocr_budget_s))
    )
    _ocr_deadline = (
        _time.monotonic() + effective_ocr_budget_s
        if effective_ocr_budget_s > 0 else None
    )
    if not force:
        cached = is_fresh(out_dir, data_path, script_path, now=decision_at)
        if cached:
            needs_complete_offline = (
                effective_ocr_budget_s == 0
                and cached.get("offline_extraction_complete") is not True
            )
            # A manifest built WITHOUT vision (e.g. the tesseract-only pre-flight gate) must be rebuilt
            # when Claude vision is now AVAILABLE (key present) and there are visual sources to UPGRADE —
            # otherwise the run would silently reuse the gate's tesseract text and never call vision. Gate
            # on _vision_configured() (not just requested) so a keyless run doesn't rebuild every time for
            # nothing (tesseract already did all it can).
            needs_vision_upgrade = _vision_configured() and not cached.get("vision_mode") and any(
                s.get("kind") == "image"
                or str(s.get("note") or "").startswith("OCR")
                or (s.get("kind") == "pdf" and s.get("status") == "fail")
                for s in cached.get("sources", []))
            if not needs_vision_upgrade and not needs_complete_offline:
                t = cached.get("totals", {})
                print(f"[extract_pool] fresh — {t.get('tabs',0)} tabs across "
                      f"{t.get('workbooks',0)} workbook(s), {t.get('extracts_written',0)} "
                      f"extract(s) already in {out_dir} (use --force to rebuild)")
                _ensure_durable_relationships(data_path, out_dir)
                if corpus_path:
                    digest = (cached.get("generation") or {}).get("digest")
                    if isinstance(digest, str):
                        frozen = _verified_generation_manifest(data_path, out_dir, digest)
                        _atomic_bytes(
                            corpus_path,
                            _read_bound_generation_artifact(out_dir, frozen, "corpus.txt"),
                        )
                    else:
                        _write_corpus(out_dir, data_path, corpus_path, cached)
                return cached
            if needs_complete_offline:
                print("[extract_pool] replacing time-budgeted visual read with a complete offline pass")
            else:
                print("[extract_pool] upgrading image-only/scanned sources to Claude vision (was OCR/unread)")

    os.makedirs(out_dir, exist_ok=True)
    # Complete traversal is a transaction precondition.  Nothing is extracted
    # until this returns successfully, so an unreadable branch can never be
    # mistaken for an empty pool.
    inventory = _complete_pool_inventory(data_path)
    build_context = tempfile.TemporaryDirectory(prefix=".extract-build-", dir=out_dir)
    build_out_dir = build_context.name
    raw_subject = os.path.basename(os.path.realpath(os.path.abspath(data_path))) or "pool"
    if raw_subject in {".", ".."} or os.sep in raw_subject:
        raise PoolInventoryError(f"unsafe pool subject name: {raw_subject!r}")
    raw_prefix = f"raw/{raw_subject}"
    # The exact captured inputs are part of the immutable generation.  They
    # remain available to every frozen child and are hashed alongside all
    # derived artifacts instead of disappearing with a temporary snapshot.
    snapshot_pool = os.path.join(build_out_dir, *raw_prefix.split("/"))
    os.makedirs(snapshot_pool)
    used = set()
    sources = []
    # Capture the whole, already-enumerated pool before interpreting anything.
    # Provenance and payload verification below consume these same immutable
    # bytes, eliminating the old sidecar-before-snapshot race where new payload
    # bytes could inherit provenance parsed from an earlier Drive state.
    captured = _capture_pool_snapshot(data_path, inventory, snapshot_pool)
    snapshot_inventory = tuple(entry["snapshot_path"] for entry in captured)
    prov_map = _collect_sidecars(
        snapshot_pool, decision_at, inventory=snapshot_inventory,
        canonical_data_path=data_path,
    )  # external provenance sidecars (EXTERNAL_DATA.md)
    snapshot_by_rel = {entry["rel"]: entry for entry in captured}
    methodology_only_rels = set()
    rel = None

    def _add(row):
        """Append a source row, enriched with relpath / external flag / provenance.
        Reads the loop's current `rel` late-bound, so it always tags the file being processed."""
        sources.append(_finish_row(row, rel, prov_map))

    def _reject_extracted_methodology(base, ext, text, provenance):
        """Quarantine a methodology-only source discovered by its full extracted text.

        The cheap pre-read above intentionally does not invoke vision/OCR and only
        sniffs a bounded text head.  A renamed scanned PDF/image therefore has no
        identity until its full reader returns.  This second gate runs *after* the
        format-specific reader (including OCR/vision) and *before* an extract,
        manifest-approved row, corpus entry, or citable provenance is written.
        """
        nonlocal n_fail
        reason = methodology_only_source_reason(base, text, provenance)
        if not reason:
            return False
        n_fail += 1
        methodology_only_rels.add(rel)
        error = f"{METHODOLOGY_ONLY_ERROR}: {reason}"
        prov_map[rel] = {"_integrity_error": error}
        _add({"file": base, "ext": ext, "kind": "methodology-only",
              "status": "fail", "error": error})
        return True

    n_workbooks = n_tabs = n_written = n_fail = 0

    for entry in captured:
        source_path = entry["path"]
        base = os.path.basename(source_path)
        if (base.startswith(".")
                or os.path.dirname(source_path).rstrip("/").endswith("_pool_extracts")):
            continue
        if _is_sidecar(base):
            continue  # metadata, folded into its document's row via prov_map — never a source
        rel = entry["rel"].replace("/", os.sep)
        ext = base.lower().rsplit(".", 1)[-1] if "." in base else ""
        stem = _sanitize(base.rsplit(".", 1)[0] if "." in base else base, "file")
        # Every parser sees only the immutable byte snapshot captured before
        # extraction began; no child is allowed to wander back to live Drive.
        p = snapshot_by_rel[entry["rel"]]["snapshot_path"]

        provenance = prov_map.get(rel) or {}
        if provenance.get("_projection_state") == "superseded":
            _add({
                "file": base, "ext": ext or "(none)", "kind": "external",
                "status": "skipped",
                "error": "superseded connector projection; retained only for audit/replay",
            })
            continue
        if provenance.get("_projection_state") == "disputed":
            _add({
                "file": base, "ext": ext or "(none)", "kind": "external",
                "status": "skipped",
                "error": "disputed connector series (providers disagree); excluded from current evidence",
            })
            continue
        if provenance.get("_projection_state") == "expired":
            _add({
                "file": base, "ext": ext or "(none)", "kind": "external",
                "status": "skipped",
                "error": "expired connector projection; excluded from current evidence",
            })
            continue
        integrity_error = provenance.get("_integrity_error")
        methodology_reason = methodology_only_source_reason(base, provenance=provenance)
        if not methodology_reason and not integrity_error:
            methodology_reason = methodology_only_source_reason(
                base, sniff_text(p, METHODOLOGY_ONLY_SNIFF_CHARS), provenance,
            )
        if methodology_reason:
            n_fail += 1
            methodology_only_rels.add(rel)
            error = f"{METHODOLOGY_ONLY_ERROR}: {methodology_reason}"
            # Never attach the report's provider/title/sidecar as citable runtime
            # provenance.  `_finish_row` maps this internal integrity marker to
            # the same small unusable-provenance object as any quarantined input.
            prov_map[rel] = {"_integrity_error": error}
            _add({"file": base, "ext": ext, "kind": "methodology-only",
                  "status": "fail", "error": error})
            continue
        if integrity_error:
            n_fail += 1
            _add({"file": base, "ext": ext, "kind": "external", "status": "fail",
                  "error": integrity_error})
            continue

        # Route by SNIFFED content, not extension — Capital IQ mislabels files
        # (a `.doc` that is MHTML, a `.rtf` that is binary Word, an `.xls` that is HTML).
        fmt = sniff_format(p)

        # ---- workbooks: OLE2 (BIFF .xls) or zip (.xlsx). The same OLE2/zip container can
        #      hold a Word .doc/.docx, so a non-workbook parse error falls through to a
        #      document read instead of failing. A missing reader lib fails LOUDLY. ----
        if fmt in ("ole2", "zip"):
            n_workbooks += 1
            wbext = "xls" if fmt == "ole2" else "xlsx"
            try:
                sheets = read_workbook(p, wbext)
            except ModuleNotFoundError as e:  # xlrd/openpyxl absent — never silent garbage
                n_workbooks -= 1
                n_fail += 1
                lib = "xlrd" if fmt == "ole2" else "openpyxl"
                _add({"file": base, "ext": ext, "kind": "workbook",
                                "status": "missing-dependency",
                                "error": f"{type(e).__name__}: {e} — run "
                                         f"`.claude/tools/setup-tools.sh` (installs {lib})",
                                "sheets": []})
                continue
            except Exception as e:  # noqa — an OLE2/zip that is NOT a sheet is a Word doc
                n_workbooks -= 1
                txt, derr = _read_doc(p)
                if txt:
                    if _reject_extracted_methodology(base, ext, txt, provenance):
                        continue
                    out_name = _unique(used, stem) + ".txt"
                    _atomic_text(os.path.join(build_out_dir, out_name), txt)
                    n_written += 1
                    _add({"file": base, "ext": ext, "kind": "document",
                                    "status": "ok",
                                    "note": f"{fmt} container, not a workbook — read as document",
                                    "extract": out_name, "chars": len(txt)})
                else:
                    n_fail += 1
                    _add({"file": base, "ext": ext, "kind": "workbook",
                                    "status": "fail",
                                    "error": f"not a workbook ({type(e).__name__}); doc fallback: {derr}",
                                    "sheets": []})
                continue
            # Test the complete workbook representation, not the bounded
            # pre-sniff.  A title on a later tab must not evade the boundary.
            workbook_text = "\n".join(
                _tab_text(base, nm, i, len(sheets), rows, nc)
                for i, (nm, _nr, nc, rows) in enumerate(sheets, 1)
            )
            if _reject_extracted_methodology(base, ext, workbook_text, provenance):
                n_workbooks -= 1
                continue
            sheet_meta = []
            for i, (nm, nr, nc, rows) in enumerate(sheets, 1):
                out_name = _unique(used, f"{stem}__{_sanitize(nm, f'sheet{i}')}") + ".txt"
                _atomic_text(
                    os.path.join(build_out_dir, out_name),
                    _tab_text(base, nm, i, len(sheets), rows, nc),
                )
                n_written += 1
                n_tabs += 1
                sheet_meta.append({"name": nm, "rows": nr, "cols": nc,
                                   "cells": _nonempty_cells(rows), "units": _detect_units(rows),
                                   "extract": out_name})
            _add({"file": base, "ext": ext, "kind": "workbook",
                            "status": "ok", "sheets": sheet_meta})

        # ---- document / text formats, again by sniffed content ----
        elif fmt in ("pdf", "mime", "rtf", "html"):
            note = None
            if fmt == "pdf":
                txt, err, note = _read_pdf(p); kind = "pdf"  # note set when text came from OCR
            elif fmt == "mime":
                txt, err = _read_mhtml(p); kind = "mhtml"
            elif fmt == "rtf":
                txt, err = _read_rtf(p); kind = "rtf"
            else:
                txt, err = _read_html(p); kind = "html"
            if txt:
                if _reject_extracted_methodology(base, ext, txt, provenance):
                    continue
                out_name = _unique(used, stem) + ".txt"
                _atomic_text(os.path.join(build_out_dir, out_name), txt)
                n_written += 1
                src = {"file": base, "ext": ext, "kind": kind, "status": "ok",
                       "extract": out_name, "chars": len(txt)}
                if note:
                    src["note"] = note  # e.g. "OCR'd (was image-only scan)"
                _add(src)
            else:
                n_fail += 1
                # A valid PDF that yields no text is image-only/scanned. Keep any SPECIFIC visual-read
                # reason (vision/OCR/timeout/key hint); only synthesize the generic needs-OCR row for the
                # bare low-level "no text layer" case, so the fix is one command away.
                if kind == "pdf":
                    low = (err or "").lower()
                    if not any(w in low for w in ("ocr", "vision", "anthropic_api_key", "tesseract",
                                                   "deferred", "timed out", "rasterise")):
                        err = ("image-only/scanned PDF — no extractable text layer (needs OCR; "
                               "run .claude/tools/setup-tools.sh to install tesseract, then re-check)")
                _add({"file": base, "ext": ext, "kind": kind,
                                "status": "fail", "error": err})

        elif fmt == "image" or ext in IMAGE_EXTS:
            # standalone image (screenshot / photographed table / chart) — read via Claude vision or OCR
            txt, note, err = _read_image_file(p); kind = "image"
            if txt:
                if _reject_extracted_methodology(base, ext or "(none)", txt, provenance):
                    continue
                out_name = _unique(used, stem) + ".txt"
                _atomic_text(os.path.join(build_out_dir, out_name), txt)
                n_written += 1
                src = {"file": base, "ext": ext or "(none)", "kind": kind, "status": "ok",
                       "extract": out_name, "chars": len(txt)}
                if note:
                    src["note"] = note  # e.g. "read via Claude vision"
                _add(src)
            else:
                n_fail += 1
                _add({"file": base, "ext": ext or "(none)", "kind": kind,
                                "status": "fail", "error": err or _no_reader_note("image")})

        elif ext in POINTER_EXTS:  # pointer stubs are JSON text — match by extension
            _add({"file": base, "ext": ext, "kind": "gdrive-pointer",
                            "status": "skipped",
                            "error": "Google Drive pointer stub — open in browser; no local content"})

        elif fmt == "text":
            # `sniff_text` is deliberately bounded.  Read the complete original
            # before approving an in-place source so an identity beyond that
            # bound cannot enter the corpus.
            full_text = open(p, encoding="utf-8", errors="ignore").read()
            if _reject_extracted_methodology(base, ext or "(none)", full_text, provenance):
                continue
            _add({"file": base, "ext": ext or "(none)", "kind": "text",
                            "status": "in-place", "extract": "(original)"})

        elif fmt == "empty":
            n_fail += 1
            _add({"file": base, "ext": ext or "(none)", "kind": "other",
                            "status": "fail", "error": "empty file"})

        else:  # unknown binary — last-ditch document conversion, else surface honestly
            txt, derr = _read_doc(p)
            if txt:
                if _reject_extracted_methodology(base, ext or "(none)", txt, provenance):
                    continue
                out_name = _unique(used, stem) + ".txt"
                _atomic_text(os.path.join(build_out_dir, out_name), txt)
                n_written += 1
                _add({"file": base, "ext": ext or "(none)", "kind": "document",
                                "status": "ok", "note": "unrecognized binary — read via textutil",
                                "extract": out_name, "chars": len(txt)})
            else:
                _add({"file": base, "ext": ext or "(none)", "kind": "other",
                                "status": "skipped", "error": f"unrecognized format ({fmt}); {derr}"})

    # Everything below is still private.  Derived artifacts and the corpus are
    # built from the immutable snapshot, never from a second live-pool walk.
    entities = _entities_from_paths(
        snapshot_pool, tuple(entry["snapshot_path"] for entry in captured),
    )
    manifest = {
        "data_path": os.path.abspath(data_path),
        "out_dir": os.path.abspath(out_dir),
        "generated_by": "extract_pool.py",
        "vision_mode": bool(_vision_on and _vision_configured()),  # were visual sources read by Claude vision?
        # True means every local reader was allowed to finish or reach its own
        # explicit file-level outcome; no pool-wide deadline deferred work.
        "offline_extraction_complete": _ocr_deadline is None,
        "sources": sources,
        "entities": entities,
        "totals": {"sources": len(sources), "workbooks": n_workbooks,
                   "tabs": n_tabs, "extracts_written": n_written, "failures": n_fail},
    }
    _write_ciq_facts(snapshot_pool, build_out_dir)
    _write_relationships(snapshot_pool, build_out_dir)
    # Always retain a frozen corpus so a later supervisor-bound child can reuse
    # this generation without reopening in-place text files on Drive.
    _write_corpus(
        build_out_dir, snapshot_pool, os.path.join(build_out_dir, "corpus.txt"),
        manifest, methodology_only_rels,
    )
    artifacts = _hash_generation_artifacts(build_out_dir)
    local_sources = _localize_original_sources(sources, raw_prefix)
    binding = _generation_binding_payload(
        data_path, out_dir, captured, local_sources, entities, artifacts,
        manifest["vision_mode"], manifest["offline_extraction_complete"], raw_prefix,
    )
    generation_binding_json = _generation_binding_json(binding)
    generation_digest = _generation_digest_from_json(generation_binding_json)
    manifest["sources"] = _prefix_generation_sources(local_sources, generation_digest)
    manifest["generation"] = {
        "schema_version": "pool-generation/v2",
        "digest": generation_digest,
        "binding_json": generation_binding_json,
        "raw_prefix": raw_prefix,
        "inputs": binding["inputs"],
        "artifacts": artifacts,
    }
    # Compute every non-pool guard before the final live comparison.  Then
    # replace only the cheap pool metadata rows with the just-verified values,
    # keeping revalidation immediately adjacent to publication.
    manifest["cache_guard"] = _cache_guard_snapshot(
        data_path, out_dir, manifest, captured=captured, extract_root=build_out_dir,
    )
    verified = _revalidate_pool_inventory(data_path, captured)
    manifest["cache_guard"]["pool_inputs"] = {
        entry["rel"]: {"size": entry["size"], "mtime_ns": entry["mtime_ns"]}
        for entry in verified
    }

    generation_parent = os.path.join(out_dir, _GENERATION_DIR_NAME)
    generation_dir = os.path.join(generation_parent, generation_digest)
    os.makedirs(generation_parent, exist_ok=True)
    _atomic_json(os.path.join(build_out_dir, "manifest.json"), manifest)
    if os.path.exists(generation_dir):
        # Content-addressed generations are immutable.  A matching one may
        # already exist after a forced deterministic rebuild; verify it and
        # discard the private duplicate rather than overwriting files a reader
        # may hold through an older manifest.
        # Recover safely if a process died after atomic rename but before the
        # chmod seal: verify content without trusting it, seal, then verify all
        # hashes and modes again. A normal forced deterministic rebuild takes
        # this same path and never rewrites the existing generation.
        _verified_generation_manifest(
            data_path, out_dir, generation_digest, require_read_only=False,
        )
        _seal_generation_tree(generation_dir)
        _verified_generation_manifest(data_path, out_dir, generation_digest)
        build_context.cleanup()
    else:
        os.replace(build_out_dir, generation_dir)
        _seal_generation_tree(generation_dir)
        _verified_generation_manifest(data_path, out_dir, generation_digest)
        if os.name != "nt":
            parent_fd = os.open(generation_parent, os.O_RDONLY)
            try:
                os.fsync(parent_fd)
            finally:
                os.close(parent_fd)

    # Prepare every mutable compatibility projection first and publish root
    # manifest.json LAST as the transaction's single commit marker.  Both old
    # and new pointers reference immutable generation directories, so a reader
    # that loaded the old pointer can never observe new extract bytes.
    _atomic_text(os.path.join(out_dir, "manifest.md"), _manifest_md(manifest))

    # Fixed-name compatibility artifacts are projections of the immutable
    # generation, not its commit marker.  Consumers that need transactional
    # binding use manifest.generation/artifacts; legacy readers keep working.
    for name in ("ciq_facts.json", "relationships.json"):
        if name in manifest["generation"]["artifacts"]:
            _atomic_bytes(
                os.path.join(out_dir, name),
                _read_bound_generation_artifact(out_dir, manifest, name),
            )
    if (os.path.basename(os.path.normpath(out_dir)) == "_pool_extracts"
            and "relationships.json" in manifest["generation"]["artifacts"]):
        _atomic_bytes(
            os.path.join(os.path.dirname(out_dir), "relationships.json"),
            _read_bound_generation_artifact(out_dir, manifest, "relationships.json"),
        )
    if corpus_path:
        _atomic_bytes(
            corpus_path,
            _read_bound_generation_artifact(out_dir, manifest, "corpus.txt"),
        )
    _atomic_json(os.path.join(out_dir, "manifest.json"), manifest)

    print(f"[extract_pool] {len(sources)} source(s) | {n_workbooks} workbook(s) "
          f"-> {n_tabs} tab(s) | {n_written} extract(s) written | {n_fail} failure(s)")
    print(f"[extract_pool] manifest: {os.path.join(out_dir, 'manifest.md')}")
    print(f"[extract_pool] generation: {generation_digest}")
    return manifest


def _write_ciq_facts(data_path, out_dir):
    """B3 generation hook: emit ciq_facts.json (deterministic, source-bound CIQ numbers) next to
    manifest.json. Wired HERE so the facts are produced automatically wherever the pool is built (every
    00-triage's full extract), with no prompt change. Fully best-effort — a facts-layer failure (a missing
    lib, a malformed workbook, a bug in a reader) must NEVER fail the pool extract. NO agent consumes it
    yet; this only PUBLISHES the facts for verify-evidence + (later) the specialists to cite."""
    try:
        from pathlib import Path as _Path
        from ciq_facts import build_facts  # same tools dir (on sys.path when run as a script)
        ticker = os.path.basename(os.path.normpath(data_path))
        facts = build_facts(_Path(data_path), ticker)
        _atomic_json(os.path.join(out_dir, "ciq_facts.json"), facts)
        n_present = sum(1 for v in facts.get("facts", {}).values() if v.get("status") == "present")
        print(f"[extract_pool] ciq_facts.json: {n_present} source-bound fact(s) "
              f"({facts.get('concepts_resolved', 0)} CIQ concepts resolved)")
    except Exception as e:  # noqa: BLE001 — best-effort; a facts failure must never fail extraction
        print(f"[extract_pool] ciq_facts: skipped ({type(e).__name__}: {e})")


def _relationship_artifact_paths(out_dir):
    """Cache path plus the committed run-root artifact when this is a canonical run extract."""
    cache = os.path.join(out_dir, "relationships.json")
    paths = [cache]
    if os.path.basename(os.path.normpath(out_dir)) == "_pool_extracts":
        paths.append(os.path.join(os.path.dirname(os.path.normpath(out_dir)), "relationships.json"))
    return paths


def _atomic_bytes(path, value):
    """Durably replace one cache artifact without exposing partial bytes."""
    directory = os.path.dirname(path) or "."
    os.makedirs(directory, exist_ok=True)
    old_mode = 0o644
    try:
        existing = os.stat(path, follow_symlinks=False)
        if stat.S_ISREG(existing.st_mode):
            old_mode = stat.S_IMODE(existing.st_mode)
    except FileNotFoundError:
        pass
    fd, tmp = tempfile.mkstemp(
        prefix=f".{os.path.basename(path)}.", suffix=".pending", dir=directory,
    )
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(value)
            os.fchmod(handle.fileno(), old_mode)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(tmp, path)
        tmp = None
        # Windows cannot open/fsync directories; file fsync + os.replace above remain portable.
        if os.name != "nt":
            dir_fd = os.open(directory, os.O_RDONLY)
            try:
                os.fsync(dir_fd)
            finally:
                os.close(dir_fd)
    finally:
        if tmp and os.path.exists(tmp):
            os.unlink(tmp)


def _atomic_text(path, value):
    _atomic_bytes(path, value.encode("utf-8"))


def _atomic_json(path, value):
    _atomic_text(path, json.dumps(value, indent=2, default=str) + "\n")


def _ensure_durable_relationships(data_path, out_dir):
    """Publish an older cached graph into the run root without forcing a full pool re-extract."""
    cache, *durable = _relationship_artifact_paths(out_dir)
    if not durable:
        return
    try:
        with open(cache, encoding="utf-8") as handle:
            graph = json.load(handle)
        if os.path.exists(durable[0]):
            with open(durable[0], encoding="utf-8") as handle:
                if json.load(handle) == graph:
                    return
        _atomic_json(durable[0], graph)
        print(f"[extract_pool] relationships.json: published durable run artifact {durable[0]}")
    except FileNotFoundError:
        _write_relationships(data_path, out_dir)
    except Exception as e:  # noqa: BLE001 — keep extraction cache hits best-effort
        print(f"[extract_pool] relationships durable publish: skipped ({type(e).__name__}: {e})")


def _write_relationships(data_path, out_dir):
    """Emit relationships.json — the deterministic supply-chain graph parsed from any CIQ
    Suppliers / Customers export in the pool, next to manifest.json and ciq_facts.json.

    Wired here for the same reason the facts sidecar is: the graph is then produced automatically
    wherever the pool is built (every module's 00-triage full extract), with no prompt change and no
    LLM spend. A pool with no relationship export writes a valid, empty graph — that is a real answer
    ("no supplier/customer list has been provided"), not a failure, and downstream readers need to be
    able to tell the two apart. Fully best-effort: a parse failure must NEVER fail the pool extract."""
    try:
        from pathlib import Path as _Path
        from relationship_graph import build_graph  # same tools dir (on sys.path when run as a script)
        ticker = os.path.basename(os.path.normpath(data_path))
        graph = build_graph(_Path(data_path), ticker)
        for output in _relationship_artifact_paths(out_dir):
            _atomic_json(output, graph)
        c = graph.get("concentration", {})
        n_sheets = len(graph.get("sources", []))
        if n_sheets:
            print(f"[extract_pool] relationships.json: {n_sheets} relationship sheet(s) -> "
                  f"{c.get('relationship_rows', 0)} disclosed row(s), "
                  f"{c.get('third_party_entities', 0)} outside part(y/ies) "
                  f"({c.get('listed_third_parties', 0)} listed)")
        elif graph.get("warnings"):
            print(f"[extract_pool] relationships.json: no relationship sheet parsed; "
                  f"{len(graph['warnings'])} input warning(s) recorded")
        else:
            print("[extract_pool] relationships.json: no CIQ supplier/customer export in the pool")
    except Exception as e:  # noqa: BLE001 — best-effort; a graph failure must never fail extraction
        print(f"[extract_pool] relationships: skipped ({type(e).__name__}: {e})")


def _manifest_md(m):
    lines = ["# Pool Extraction Manifest", "",
             f"- Data pool: `{m['data_path']}`",
             f"- Extracts: `{m['out_dir']}`",
             f"- Totals: {m['totals']['workbooks']} workbook(s) → "
             f"{m['totals']['tabs']} tab(s); {m['totals']['extracts_written']} extract file(s); "
             f"{m['totals']['failures']} failure(s)", "",
             "| Source File | Type | Tab / Stream | Rows×Cols | Cells | Extract |",
             "|---|---|---|---|---|---|"]
    for s in m["sources"]:
        display = s.get("path", s["file"])  # nested (e.g. external/) files show their pool-relative path
        prov = _prov_summary(s)
        if s["kind"] == "workbook" and s.get("sheets"):
            kind = f"workbook ({prov})" if prov else "workbook"
            for sh in s["sheets"]:
                lines.append(f"| {display} | {kind} | {sh['name']} | "
                             f"{sh['rows']}×{sh['cols']} | {sh['cells']} | `{sh['extract']}` |")
        else:
            note = s.get("extract", s.get("error", "—"))
            if prov:
                note = f"{note} — {prov}"
            lines.append(f"| {display} | {s['kind']} ({s['status']}) | — | — | — | {note} |")
    return "\n".join(lines) + "\n"


def _html_xls_fallback(path):
    """Some '.xls' files are actually HTML tables. textutil can flatten them to text."""
    txt, _ = _read_rtf(path)  # textutil -convert txt also reads HTML
    return txt or ""


def _write_corpus(out_dir, data_path, corpus_path, manifest, excluded_rels=None):
    """Concatenate only manifest-approved extracts and original text sources.

    Restricting the corpus to the current manifest also prevents a stale extract
    from an earlier run from surviving after its source is quarantined.  Explicit
    exclusions are the methodology-only files rejected during this same pass.
    """
    parts = []
    excluded = {str(rel).replace(os.sep, "/") for rel in (excluded_rels or set())}
    approved_extracts = set()
    approved_originals = set()
    for source in (manifest or {}).get("sources", []):
        if source.get("status") not in ("ok", "in-place"):
            continue
        source_rel = str(source.get("path") or source.get("file") or "").replace(os.sep, "/")
        extract = source.get("extract")
        if extract == "(original)":
            approved_originals.add(source_rel)
        elif isinstance(extract, str) and extract.endswith(".txt"):
            approved_extracts.add(extract)
        for sheet in source.get("sheets") or []:
            sheet_extract = sheet.get("extract") if isinstance(sheet, dict) else None
            if isinstance(sheet_extract, str) and sheet_extract.endswith(".txt"):
                approved_extracts.add(sheet_extract)

    for name in sorted(approved_extracts):
        p = os.path.join(out_dir, name)
        if not os.path.isfile(p):
            continue
        parts.append(f"\n===== EXTRACT: {os.path.basename(p)} =====\n")
        parts.append(open(p, errors="ignore").read())
    for p in iter_pool_files(data_path):
        rel = os.path.relpath(p, data_path).replace(os.sep, "/")
        if p.lower().endswith(".txt") and rel in approved_originals and rel not in excluded:
            # Read BEFORE labelling. Appending the header first meant a failed read left a
            # labelled but EMPTY section in the corpus — worse than omitting the source, because
            # verify-evidence then greps a section that exists, finds nothing in it, and reads that
            # as "the source says nothing" rather than "the source could not be read" (§20 bad
            # extraction). An unreadable source must be absent from the corpus, not silently blank.
            try:
                body = _read_pool_regular_bytes(data_path, rel).decode("utf-8", "ignore")
            except (OSError, ValueError):
                continue
            # relpath, not basename — data/T/note.txt and data/T/external/x/note.txt must be
            # labelled distinctly, or verify-evidence attributes greps to the wrong source.
            parts.append(f"\n===== SOURCE: {rel} =====\n")
            parts.append(body)
    _atomic_text(corpus_path, "".join(parts))
    print(f"[extract_pool] corpus: {corpus_path} ({sum(len(x) for x in parts)} chars)")


# ---------- readiness pre-flight (the data-readiness GATE; deterministic, no LLM) ----------
#
# Feeds the cockpit's pre-spawn gate. Summarizes the pool's readiness from the SAME extraction
# truth the pipeline uses (manifest sources[].status), plus a surface-and-confirm entity read.
# severity: 'blocker' (wrong-direction / garbage if proceeded) | 'degrade' (same-direction, weaker)
# | 'info' (note only). The launcher folds in type-based issues (no-price, annual-vs-quarterly,
# §26 module readiness) from data-status.ts — this stays extraction- and entity-truth only.

# A file COUNTS as usable when it has real content: 'ok' (extracted) OR 'in-place' (a plain text / CSV /
# MD file used as-is). Everything else (fail / missing-dependency / skipped / gdrive-pointer stub) is not.
USABLE_STATUSES = {"ok", "in-place"}

_STATUS_ISSUES = {                       # per-source NON-usable status -> (issue code, severity)
    "fail":               ("extraction_failed", "degrade"),
    "missing-dependency": ("missing_dependency", "degrade"),
    "skipped":            ("unreadable_format",  "info"),
}

_CORP_SUFFIX = (r"(?:Limited|Ltd\.?|Inc\.?|Incorporated|Corporation|Corp\.?|Company|"
                r"PLC|LLC|N\.V\.|S\.A\.|S\.p\.A\.|AG|SE|Holdings|Group)")

# Don't read a registrant name from non-filings that name PEERS: transcripts/decks (peers + people) and
# comparables/comps (a comp sheet's whole point is to list rivals). Key-developments + most CIQ data
# exports DO carry a clean "{Company} (EXCH:TICKER)" subject header (see _CIQ_HEADER), so they're allowed.
_NON_FILING_NAME = re.compile(
    r"(transcript|earnings.?call|\bcall\b|presentation|investor.?(?:deck|present)|\bdeck\b|\bppt\b|"
    r"conference|webcast|comparable|\bcomps?\b)",
    re.I)

# CIQ / data-vendor export header: "{Company} ({Exchange}:{Ticker}) > Report > Section". The SUBJECT
# company precedes the (EXCH:TICKER) tag — a high-confidence registrant signal that also works for
# filings whose cover carries their own ticker. Used FIRST + only in the header region (so a deep-prose
# peer mention can't match), which lets data exports contribute their clean entity.
# The name and its (EXCH:TICKER) tag must sit on the SAME line, and the name may not begin
# mid-word. `\s*` (which matches newlines) let this span a line break, and the {2,80} bound then
# slid the start forward until the span fit — turning an Indian cover letter's two-column addressee
# block ("BSE Limited        National Stock Exchange of India Limited\n(BSE: 542726)") into the
# left-truncated garbage entities "ited National Stock Exchange of India Limited", "ed National
# Stock Exchange…", "d National Stock Exchange…". Three fake companies from one line, which then
# read as a mixed-entity pool and blocked the run.
_CIQ_HEADER = re.compile(
    r"(?<![A-Za-z0-9])([A-Za-z][^\S\r\n\w]*[\w&.,'’\- ]{2,80}?)[^\S\r\n]*"
    r"\(\s*[A-Za-z][\w.]*\s*:\s*[A-Za-z0-9.\-]+\s*\)")

# Exchanges, depositories and regulators are the ADDRESSEE of a filing cover letter, never its
# registrant. Every Indian filing opens "To, BSE Limited / National Stock Exchange of India
# Limited", so without this the exchange is read as a second company in the pool and the
# mixed-entity gate blocks a perfectly clean pool (CLAUDE.md §27: an Indian company is the
# default-likely case). Matched on the normalized key, so punctuation/suffix variants collapse.
_ADDRESSEE_ENTITY_KEYS = frozenset({
    "nationalstockexchangeindia", "nationalstockexchange", "bombaystockexchange",
    "nationalsecuritiesdepository", "centraldepositoryservicesindia",
    "centraldepositoryservices", "metropolitanstockexchangeindia",
    "securitiesexchangeboardindia", "newyorkstockexchange", "londonstockexchange",
    "nasdaqstockmarket", "singaporeexchangesecuritiestrading",
})
# Cover-letter framing that proves the head is an ADDRESSEE block rather than a title page. The
# stop-list only applies when one of these is present, so an exchange that is itself the subject of
# the run (BSE Limited and NSE are both listed companies) is still read from its own title page.
# Deliberately NARROW. Two earlier drafts of this were far too broad: `\bTrading\s+Symbol\b` sits on
# every post-2019 US 10-K/10-Q/8-K cover page, and a bare `^\s*To[,:]` under re.M matched any wrapped
# prose line beginning "to, ". Both armed the addressee branch on documents that are not cover
# letters at all. The "To," cue must therefore be a STANDALONE addressee line, and the US
# cover-page cue is gone entirely.
_ADDRESSEE_CONTEXT = re.compile(
    r"^\s*To[,:]\s*$|\bListing\s+(?:Department|Compliance)\b|\bDear\s+Sir\b|\bScrip\s+Code\b"
    r"|\bBSE\s+Scrip\b", re.I | re.M)
# Indian cover letters sign off "Yours faithfully, / For <REGISTRANT>" — a high-precision POSITIVE
# signal for the true issuer. The sign-off cue is REQUIRED: `For <Capitalised words> Group` on its
# own matches ordinary boilerplate ("For Further Information Contact Investor Relations Group"),
# because _CORP_SUFFIX contains plain English words (Group, Company, Holdings). Without the cue this
# fabricated a company name and — being checked first — overrode the correct registrant on US and
# data-vendor documents. It is also checked AFTER the two high-confidence patterns now, so it can
# only ever fill a gap, never override a header that actually names the registrant.
_SIGNATORY_FOR = re.compile(
    r"(?:Yours\s+(?:faithfully|sincerely|truly)|Thanking\s+you)[^\n]*\n\s*"
    r"(?:For|for)\s+([A-Z][A-Za-z0-9&.,'’\- ]{2,78}?" + _CORP_SUFFIX + r")\s*$",
    re.I | re.M)


def _clean_entity(s):
    return re.sub(r"\s+", " ", s or "").strip(" .,:-\t–—")


def _norm_entity(s):
    """Normalize an entity name for comparison: lowercase, drop corporate suffixes + punctuation."""
    s = re.sub(r"[^a-z0-9 ]", " ", (s or "").lower())
    s = re.sub(r"\b(limited|ltd|inc|incorporated|corporation|corp|company|plc|llc|nv|sa|spa|"
               r"ag|se|holdings|group|the|formerly|erstwhile)\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# Tearsheet FIELD LABELS / boilerplate that get read as fake names (a CIQ profile's "Company Type:
# Public Company", "Named by Company", "Country/Region of Incorporation"), plus generic words that
# do NOT make two names "related". Two roles: (1) a candidate whose EVERY normalized token is in
# here is a label, not a company (_looks_like_entity rejects it); (2) a token from here is not
# DISTINCTIVE, so two unrelated ".com" firms sharing only "com" still count as a conflict, not a
# soft "related" degrade. (Corporate-form words are already stripped by _norm_entity, so they never
# reach this set as tokens.)
_GENERIC_ENTITY_TOKENS = {
    "com", "co", "www", "type", "public", "private", "listed", "entity", "named", "by", "of",
    "country", "region", "incorporation", "sector", "industry", "address", "website",
    "currency", "employees", "status", "primary",
}


def _entity_key(s):
    """Punctuation/space-INSENSITIVE canonical key for an entity name: _norm_entity with spaces
    removed. 'Amazon.com, Inc', 'AMAZON.COM, INC', 'Amazon com Inc' and 'Amazoncom Inc' all
    collapse to 'amazoncom' — so filesystem-sanitized filenames (which drop '.'/',') cluster with
    the filing's own header spelling instead of reading as a different company."""
    return _norm_entity(s).replace(" ", "")


def _bounded_lev(a, b, max_d=2):
    """True Levenshtein (edit) distance between two short strings, capped: returns max_d+1 as soon
    as the whole current DP row exceeds max_d. Edit distance — not difflib's Ratcliff-Obershelp
    ratio — is the right model for OCR glyph slips ('rn'->'m' is a clean 2-edit) and caps cleanly.
    Entity keys are short, so this stays effectively O(len)."""
    if abs(len(a) - len(b)) > max_d:
        return max_d + 1
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[-1] + 1, prev[j - 1] + (ca != cb)))
        if min(cur) > max_d:
            return max_d + 1
        prev = cur
    return prev[-1]


def _same_entity(k1, k2):
    """True if two canonical keys are the SAME company: identical, or a tight OCR/typo apart (both
    reasonably long, near-equal length, edit distance <= 2). Deliberately NARROW — a false 'these
    are the same' hides real contamination, which the rejector doctrine (CLAUDE.md §24) treats as
    worse than a false block. min-length 6 (not 5) so 'apple'/'ample' (1 edit) never fuse, while
    the intended 'amazoncom'/'amazoncorn' (len 9-10) still do."""
    if k1 == k2:
        return True
    if min(len(k1), len(k2)) < 6 or abs(len(k1) - len(k2)) > 2:
        return False
    return _bounded_lev(k1, k2, 2) <= 2


# Lead-in words no registrant name starts with. A cover page or press release carries capitalised
# boilerplate that otherwise satisfies every other test — "For Further Information Contact Investor
# Relations Group" ends in a _CORP_SUFFIX ("Group"), is short, and is not mostly lowercase, so it was
# accepted as a company and counted toward the mixed-entity gate. Two files carrying it clear
# _MIN_FILES_FOR_CONFLICT and block a clean pool on a company that does not exist.
_ENTITY_LEAD_IN_WORDS = {"for", "to", "about", "contact", "attn", "attention", "re", "by", "from",
                         "dear", "subject", "ref", "reference"}


def _looks_like_entity(name):
    """A registrant name LOOKS like a name, not prose or a bare suffix word: it has a real (non-suffix)
    word, is short, and isn't mostly lowercase function words ("is in fact the very purpose of ...")."""
    words = (name or "").split()
    if not (1 <= len(words) <= 8):
        return False
    if words[0].strip(".,:").lower() in _ENTITY_LEAD_IN_WORDS:
        return False                            # boilerplate lead-in, not a registrant name
    norm = _norm_entity(name)
    if not norm:                                # nothing but corporate suffixes ("Company", "The Group")
        return False
    if all(t in _GENERIC_ENTITY_TOKENS for t in norm.split()):   # a tearsheet field LABEL, not a name
        return False                            # "Named by Company", "Company Type: Public Company", ...
    lower = sum(1 for w in words if w[:1].islower())
    return lower <= len(words) // 2             # mostly-lowercase -> prose, not a name


_MIN_FILES_FOR_CONFLICT = 2   # a different company must appear in >=2 files to flag (1 = extraction noise)


def _entity_clusters(entities):
    """Group [{file,entity}] into distinct companies by a punctuation/space-INSENSITIVE key (_entity_key),
    then fold in OCR/typo-apart keys. Case / punctuation / spacing / corporate-suffix variants collapse
    ('Amazon.com, Inc' and 'Amazoncom Inc' are one company, not two), but a BUSINESS-UNIT qualifier does
    NOT — 'Tata Motors' (key 'tatamotors') and 'Tata Motors Passenger Vehicles' (key
    'tatamotorspassengervehicles') stay separate, because for a CV ticker the PV demerged entity is a
    different company (the exact gap that read 'clean' before). Returns [{key, tokens, name (fullest
    display), files:[...]}], sorted by file count desc. A file can contribute more than one name (its
    content AND its filename)."""
    groups = {}  # canonical key -> {key, name: fullest display, tokens: set, files: [...]}
    for e in entities:
        raw = e.get("entity", "")
        nm = _norm_entity(raw)
        if not nm:
            continue
        key = nm.replace(" ", "")
        g = groups.setdefault(key, {"key": key, "name": raw, "tokens": set(), "files": []})
        g["tokens"] |= set(nm.split())               # keep the REAL (spaced) tokens for the severity test
        if e["file"] not in g["files"]:
            g["files"].append(e["file"])
        if len(raw) > len(g["name"]):
            g["name"] = raw                          # keep the fullest display name
    # Fold an OCR/typo-apart key into the LARGER surviving cluster (majority-anchored: each cluster is
    # compared only to already-ACCEPTED anchors, never to an absorbed member, so it can't chain-merge
    # A~B~C when A and C are unrelated). Order-stable (dict preserves first-seen; sort is stable).
    merged = []
    for r in sorted(groups.values(), key=lambda g: -len(g["files"])):
        for m in merged:
            if _same_entity(m["key"], r["key"]):
                for f in r["files"]:
                    if f not in m["files"]:
                        m["files"].append(f)
                m["tokens"] |= r["tokens"]
                if len(r["name"]) > len(m["name"]):
                    m["name"] = r["name"]
                break
        else:
            merged.append(r)
    merged.sort(key=lambda r: -len(r["files"]))
    return merged


def _entity_conflict(entities):
    """Severity of an entity conflict among [{file,entity}], or None. A different company must appear in
    >=_MIN_FILES_FOR_CONFLICT files to flag — a single odd extraction is treated as NOISE, because diverse
    real pools throw off one-off garbage names (an owner on a holdings sheet, an exchange on a cover page, a
    model's cell headers). Among the companies that clear the threshold and differ from the majority:
      - 'blocker' if one shares NO DISTINCTIVE word with the majority — clearly unrelated, so the pool is
        contaminated with another company's files (e.g. a batch of 'Tata Motors' files in a 'Salesforce'
        pool). A shared word that is only a GENERIC token ('com' between two unrelated '.com' firms) is not
        distinctive and does NOT soften the block;
      - 'degrade' if it still shares a real word (related group / abbreviation / weak extraction)."""
    reps = _entity_clusters(entities)
    if len(reps) < 2:
        return None
    maj = reps[0]
    maj_tokens = maj["tokens"] - _GENERIC_ENTITY_TOKENS
    level = None
    for r in reps[1:]:
        if len(r["files"]) < _MIN_FILES_FOR_CONFLICT:
            continue                                 # singleton -> extraction noise, ignore
        if not (maj_tokens & (r["tokens"] - _GENERIC_ENTITY_TOKENS)):
            return "blocker"                         # no MEANINGFUL shared word -> contamination
        level = "degrade"
    return level


def _entities_disagree(names):
    """TEST-ONLY bool used by the smoke's clustering assertions (production reads use _entity_conflict +
    _entity_evidence). Same clustering as production — punctuation/space-insensitive key with a tight
    OCR/typo fuse — so it answers "is there more than one distinct company?" ignoring the file-count
    threshold (which the smoke tests separately via _entity_conflict). NOT token-subset: a business-unit
    qualifier ('Passenger Vehicles') counts as distinct, exactly as the production path treats it."""
    return len(_entity_clusters([{"file": "", "entity": n} for n in names])) > 1


def _entity_evidence(entities):
    """A FOCUSED evidence line: name the majority company and point only at the files of the companies that
    actually FLAGGED (cleared the >=2-file threshold) — not a dump of every file, and not the ignored
    singletons."""
    reps = _entity_clusters(entities)
    flagged = [r for r in reps[1:] if len(r["files"]) >= _MIN_FILES_FOR_CONFLICT]
    if not flagged:
        return "; ".join(f'{e["file"]} → {e["entity"]}' for e in entities if e.get("entity"))
    maj = reps[0]
    odd_files = [f for r in flagged for f in r["files"]]
    odd_names = " / ".join(dict.fromkeys(r["name"] for r in flagged))
    return (f'{len(maj["files"])} file(s) name "{maj["name"]}"; {len(odd_files)} look out of place '
            f'({odd_names}): {", ".join(odd_files)}')


def entity_from_header(text):
    """Best-guess the registrant / issuer name from a filing's header text — deterministic, heuristic.
    Collects candidates from a few patterns and returns the first that LOOKS like a real name (not prose
    or a bare suffix word). Returns '' if none. Surface-and-confirm: shows what the document says; the
    engine never asserts the canonical name."""
    if not text:
        return ""
    head = text[:2500]
    cands = []
    # An Indian filing cover letter is ADDRESSED to the exchanges and SIGNED by the registrant
    # ("Yours faithfully, / For IndiaMART InterMESH Limited"). When that framing is present the
    # signature is the issuer and the addressee block is not, so read the signature FIRST and
    # discard exchange/depository names entirely (see _ADDRESSEE_ENTITY_KEYS).
    addressed = bool(_ADDRESSEE_CONTEXT.search(head))
    # CIQ / data-vendor export header "{Company} (EXCH:TICKER) > ..." — checked FIRST and only in the
    # header region, so a data export contributes its clean subject (not a peer named deeper in the file).
    m = _CIQ_HEADER.search(head[:300])
    if m:
        cands.append(m.group(1))
    # SEC: the name precedes "Exact name of registrant as specified in its charter"
    m = re.search(r"([^\n]{3,90}?)\s*\n[^\n]{0,40}Exact name of (?:the )?[Rr]egistrant", head)
    if m:
        cands.append(m.group(1))
    # India / generic: "Name of the Company / Listed Entity:" — REQUIRE a real separator (colon/dash or a
    # line break) before the value, so it can't capture inline prose ("...name of the Company is in fact").
    m = re.search(r"Name of (?:the )?(?:Listed Entity|Company)\s*(?:[:\-]\s*|\n\s*)([^\n]{3,90})", head, re.I)
    if m:
        cands.append(m.group(1))
    # An Indian filing cover letter is ADDRESSED to the exchanges and SIGNED by the registrant
    # ("Yours faithfully, / For IndiaMART InterMESH Limited"). Checked only when that framing is
    # present, and only AFTER the higher-confidence header patterns above, so it fills the gap those
    # leave on a cover letter without ever overriding a document that names its registrant outright.
    if addressed:
        m = _SIGNATORY_FOR.search(head)
        if m:
            cands.append(m.group(1))
    # the first short STANDALONE line ending in a corporate suffix (cover-page title; often ALL CAPS).
    # Strip a trailing parenthetical first ("(Formerly TML ...)", "(NYSE: CRM)") so the registrant name
    # is matched, not lost to the suffix being followed by ")".
    for line in head.splitlines():
        core = re.sub(r"\s*\([^)]*\)\s*$", "", line.strip()).strip()
        if 3 <= len(core) <= 90 and len(core.split()) <= 8 and re.search(r"\b" + _CORP_SUFFIX + r"\.?\s*$", core, re.I):
            cands.append(core)
            break
    for c in cands:
        c = _clean_entity(c)
        if not _looks_like_entity(c):       # reject prose ("fact the very purpose") + bare suffixes ("Company")
            continue
        # Substring, not equality: a PDF text layer flattens the two-column addressee block into one
        # line ("BSE Limited        National Stock Exchange of India Limited"), whose key contains an
        # addressee key without equalling it. Either way it is who the letter was sent TO.
        if addressed and any(k in _entity_key(c) for k in _ADDRESSEE_ENTITY_KEYS):
            continue
        return c
    return ""


def entity_from_filename(name):
    """Best-guess the entity from the FILE NAME. CIQ exports + downloaded filings are usually named after
    the company ("Tata_Motors_Passenger_Vehicles_Limited_-_Form_Annual_Report...", "Salesforce Inc NYSE
    CRM Financials..."). This catches a wrong-entity file whose CONTENT hides it — e.g. a demerged entity
    whose cover page still carries the former parent's name. Returns the leading name up to its first
    corporate suffix, or '' (a cryptic filename yields nothing — content still covers it)."""
    base = re.sub(r"[_\-]+", " ", os.path.splitext(name)[0])
    base = re.sub(r"\s+", " ", base).strip()
    m = re.match(r"([A-Z][A-Za-z0-9&.,' ]{2,78}?\b" + _CORP_SUFFIX + r")\b", base)
    if m:
        cand = _clean_entity(m.group(1))
        if _looks_like_entity(cand):
            return cand
    return ""


def _entities_from_paths(data_path, paths):
    """Read entity signals from one already-selected inventory."""
    entities = []
    for p in paths:
        base = os.path.basename(p)
        rel = os.path.relpath(p, data_path)
        # External documents are deliberately multi-company; sidecars are
        # metadata rather than entity evidence.
        if _is_external_rel(rel) or _is_sidecar(base):
            continue
        if _NON_FILING_NAME.search(base):
            continue
        fn = entity_from_filename(base)
        if fn:
            entities.append({"file": base, "entity": fn})
        if sniff_format(p) not in ("pdf", "mime", "rtf", "html", "ole2", "zip"):
            continue
        name = entity_from_header(sniff_text(p, 2500))
        if name:
            entities.append({"file": base, "entity": name})
    return entities


def readiness_summary(data_path, out_dir, force=False):
    """Deterministic pre-flight summary for the data-readiness gate. No LLM.
    Returns {file_count, usable_count, issues[], entities[]}."""
    # The whole body sniffs files — extract_pool prints log lines, and the entity reads can invoke pypdf
    # in-process (which may print). Redirect ordinary Python stdout for direct/module callers. The CLI's
    # _emit_machine_json additionally shields file descriptor 1 because retained stream defaults/native
    # writes bypass this assignment (the real noisy-xlrd failure this two-layer contract prevents).
    _saved_stdout = sys.stdout
    sys.stdout = sys.stderr
    try:
        # Admission is provider-neutral and complete: never spend research-provider tokens, but let
        # every local reader finish (or record its exact file-level failure) before freezing this
        # generation. Later children reuse it and never perform a hidden vision/OCR upgrade.
        manifest = extract_pool(
            data_path, out_dir, force=force, vision=False, ocr_budget_s=0,
        )
        sources = manifest.get("sources", [])
        usable = [s for s in sources if s.get("status") in USABLE_STATUSES]
        issues = []
        if not sources:
            issues.append({"code": "zero_files", "severity": "blocker",
                           "message": "No files found in the data pool — nothing to analyze.",
                           "evidence": f"data path: {os.path.abspath(data_path)}"})
        elif not usable:
            issues.append({"code": "zero_usable_data", "severity": "blocker",
                           "message": "No file in the pool could be read — nothing usable to analyze.",
                           "evidence": f"{len(sources)} file(s) present, 0 readable"})
        for s in sources:
            spec = _STATUS_ISSUES.get(s.get("status"))
            if spec:
                code, sev = spec
                issues.append({"code": code, "severity": sev, "file": s.get("file"),
                               "message": f'Could not fully read "{s.get("file")}" ({s.get("status")}).',
                               "evidence": s.get("error") or s.get("status")})

        # entity read (surface-and-confirm): sniff each file's header for the registrant/subject name.
        # Allow documents (pdf/mhtml/rtf/html) AND workbooks/Word docs (ole2/zip) — CIQ data exports carry
        # a clean "{Company} (EXCH:TICKER)" subject header (entity_from_header checks it first). Still skip
        # peer-listing non-filings by name (transcripts/decks/comparables) via _NON_FILING_NAME.
        entities = manifest.get("entities")
        if not isinstance(entities, list):
            # Backward compatibility for one legacy flat generation.  New
            # manifests always carry entities from their frozen snapshot, so a
            # supervisor-bound reuse never comes back to live Drive here.
            entities = _entities_from_paths(data_path, _complete_pool_inventory(data_path))
        _evid = _entity_evidence(entities)
        _level = _entity_conflict(entities)
        if _level == "blocker":
            # Unrelated companies in one pool = contamination: wrong-company files present and/or the
            # right ones missing. High-confidence, so it's a hard stop, not a one-click.
            issues.append({"code": "entity_disagreement", "severity": "blocker",
                           "message": "Some files are for a different company — the pool mixes entities "
                                      "(and the files you want may be missing).",
                           "evidence": _evid})
        elif _level == "degrade":
            # Distinct but RELATED names (share a word) — could be a group entity / abbreviation / weak
            # extraction, where a false positive is plausible, so surface it but allow a one-click proceed.
            issues.append({"code": "entity_disagreement", "severity": "degrade",
                           "message": "Some files may be for a different (related) company — worth a "
                                      "check before you run.",
                           "evidence": _evid})

        return {
            "data_path": os.path.abspath(data_path),
            "file_count": len(sources),
            "usable_count": len(usable),
            "issues": issues,
            "entities": entities,       # the launcher confirms these against the ticker (surface-and-confirm)
            "generation_digest": (manifest.get("generation") or {}).get("digest"),
        }
    finally:
        sys.stdout = _saved_stdout


def _emit_machine_json(build_payload):
    """Emit exactly one JSON document on stdout, even when a parser is noisy.

    Reassigning ``sys.stdout`` is not sufficient: some dependencies retain the
    original stream object in a default argument, and native code can write to
    file descriptor 1 directly.  That is exactly what xlrd does for warnings on
    some real Capital IQ OLE2 workbooks.  Shield the OS descriptor for the
    entire build and write the protocol payload through a duplicate of the
    original descriptor.  Descriptor 1 intentionally remains pointed at
    stderr until process exit, so a late flush/atexit warning cannot appear
    after the JSON document either.
    """
    try:
        stdout_fd = sys.stdout.fileno()
        stderr_fd = sys.stderr.fileno()
    except (AttributeError, ValueError, OSError):
        # Import-level/unit callers may replace stdout with StringIO.  Keep a
        # safe Python-stream fallback; the production CLI always takes the FD
        # path above.
        protocol_stdout = sys.stdout
        sys.stdout = sys.stderr
        try:
            payload = build_payload()
        finally:
            sys.stdout = protocol_stdout
        protocol_stdout.write(json.dumps(payload) + "\n")
        protocol_stdout.flush()
        return

    protocol_fd = os.dup(stdout_fd)
    try:
        sys.stdout.flush()
        os.dup2(stderr_fd, stdout_fd)
        payload = build_payload()
        # Flush every Python-level diagnostic while fd 1 still points at
        # stderr.  os.write below bypasses the redirected TextIO wrapper.
        sys.stdout.flush()
        encoded = (json.dumps(payload) + "\n").encode("utf-8")
        view = memoryview(encoded)
        while view:
            view = view[os.write(protocol_fd, view):]
    finally:
        os.close(protocol_fd)


def main(argv):
    _ensure_deps()  # re-exec under .claude/tools/.venv if xlrd/openpyxl are missing here
    if "--readiness-json" in argv:
        i = argv.index("--readiness-json")
        dp = argv[i + 1]
        rest = [a for a in argv[i + 2:] if not a.startswith("--")]
        od = rest[0] if rest else tempfile.mkdtemp(prefix="readiness_")
        _emit_machine_json(lambda: readiness_summary(dp, od, force=("--force" in argv)))
        return 0
    if "--list-json" in argv:
        i = argv.index("--list-json")
        f = argv[i + 1]
        _emit_machine_json(lambda: list_file(f))
        return 0
    if "--text" in argv:
        f = argv[argv.index("--text") + 1]
        mx = int(argv[argv.index("--max-chars") + 1]) if "--max-chars" in argv else 16000
        sys.stdout.write(sniff_text(f, mx))
        return 0
    args = [a for a in argv if not a.startswith("--")]
    if len(args) < 2:
        print(__doc__)
        return 2
    data_path, out_dir = args[0], args[1]
    corpus = None
    if "--corpus" in argv:
        corpus = argv[argv.index("--corpus") + 1]
    # vision defaults to the env (EXTRACT_VISION); --no-vision / --vision force it for tests + tooling.
    vision = None
    if "--no-vision" in argv:
        vision = False
    elif "--vision" in argv:
        vision = True
    extract_pool(data_path, out_dir, force=("--force" in argv), corpus_path=corpus, vision=vision)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
